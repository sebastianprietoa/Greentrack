from flask import Flask, render_template, request, redirect, flash, session, send_file, jsonify, url_for
import psycopg2
import psycopg2.extras
from datetime import datetime
import pandas as pd
import hashlib
import os
import io
import json
import tempfile
from extractor_sinader import extract_sinader_data
from extractor_sidrep import extract_sidrep_data, clasificar_defra
from auth_utils import hash_password, verify_password
from routes import register_blueprints

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "clave_segura_provisional")
register_blueprints(app)

@app.context_processor
def inject_pendientes_count():
    if session.get('es_admin') == 1:
        try:
            conn = get_db()
            cur = conn.cursor()
            cur.execute("SELECT COUNT(*) FROM pending_pdf_uploads WHERE estado = 'pendiente'")
            total = cur.fetchone()[0] or 0
            cur.execute("SELECT COUNT(*) FROM tickets WHERE estado = 'Abierto'")
            tickets_abiertos = cur.fetchone()[0] or 0
            conn.close()
            return {'total_pendientes': total, 'tickets_abiertos': tickets_abiertos}
        except Exception:
            return {'total_pendientes': 0, 'tickets_abiertos': 0}
    if session.get('user_id') and session.get('es_admin') == 0:
        try:
            conn = get_db()
            cur = conn.cursor()
            cur.execute("SELECT COUNT(*) FROM tickets WHERE empresa = %s AND estado != 'Cerrado'", (session.get('empresa'),))
            mis_tickets = cur.fetchone()[0] or 0
            conn.close()
            return {'mis_tickets_abiertos': mis_tickets}
        except Exception:
            return {'mis_tickets_abiertos': 0}
    return {}

# Configuración de Base de Datos para Producción (Railway) o Local
DATABASE_URL = os.getenv("DATABASE_URL")

def get_db():
    # 1. Obtenemos la variable secreta de Railway
    db_url = os.getenv("DATABASE_URL")
    
    # 2. Si Railway no está mandando la variable, esto evitará que el sistema colapse en silencio
    if not db_url:
        raise ValueError("¡Error CRÍTICO: No se encontró la variable DATABASE_URL en Railway!")
    
    # 3. Corrección automática: Railway a veces entrega la URL como "postgres://", 
    # pero Python exige que diga "postgresql://". Esto lo arregla solo.
    if db_url.startswith("postgres://"):
        db_url = db_url.replace("postgres://", "postgresql://", 1)
        
    # 4. Conectamos forzando el modo seguro (SSL), que Railway a veces exige
    return psycopg2.connect(db_url, sslmode='require')
# Filtro para números con puntos y comas
@app.template_filter('formato_cl')
def formato_cl(value):
    try:
        return "{:,.2f}".format(float(value)).replace(',', 'X').replace('.', ',').replace('X', '.')
    except:
        return value

# Filtro compacto para KPIs: muestra K o M cuando el número es grande
@app.template_filter('formato_kpi')
def formato_kpi(value):
    try:
        v = float(value)
        if v >= 1_000_000:
            return "{:,.2f} M".format(v / 1_000_000).replace(',', 'X').replace('.', ',').replace('X', '.')
        elif v >= 10_000:
            return "{:,.1f} K".format(v / 1_000).replace(',', 'X').replace('.', ',').replace('X', '.')
        else:
            return "{:,.2f}".format(v).replace(',', 'X').replace('.', ',').replace('X', '.')
    except:
        return value

# Filtro para mostrar Mes y Año
@app.template_filter('mes_anio')
def mes_anio(fecha_str):
    try:
        meses = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
        fecha_obj = datetime.strptime(fecha_str[:10], "%Y-%m-%d")
        return f"{meses[fecha_obj.month]} {fecha_obj.year}"
    except:
        return fecha_str

# Factores de emisión base
FACTORES = {
    "Electricidad": {"kWh": 0.233},
    "Transporte": {"km": 0.192},
    "Gas": {"m3": 2.0},
    "Combustible móvil": {"L": 2.68},
    "Residuos": {"kg": 0.45},
    "Agua": {"m3": 0.34}
}

def init_db():
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS usuarios (
        id SERIAL PRIMARY KEY,
        empresa TEXT UNIQUE NOT NULL,
        email TEXT UNIQUE NOT NULL,
        password TEXT NOT NULL,
        contacto TEXT,
        telefono TEXT,
        direccion TEXT,
        rut TEXT,
        es_admin INTEGER DEFAULT 0,
        fecha_registro TEXT,
        anio_default TEXT
    )
    """)
    cursor.execute("ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS anio_default TEXT")
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS registros (
        id SERIAL PRIMARY KEY,
        fecha TEXT,
        empresa TEXT,
        area TEXT,
        alcance TEXT,
        fuente TEXT,
        subfuente TEXT,
        categoria TEXT,
        actividad TEXT,
        identificador TEXT,
        detalle TEXT,
        unidad TEXT,
        cantidad REAL,
        costo REAL,
        moneda TEXT,
        factor REAL,
        emision REAL
    )
    ''')
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS configuracion (
        id SERIAL PRIMARY KEY,
        empresa TEXT UNIQUE,
        info_por_unidad INTEGER,
        vehiculos INTEGER DEFAULT 0,
        combustible_colaboradores INTEGER,
        tarjeta_combustible INTEGER,
        FOREIGN KEY (empresa) REFERENCES usuarios(empresa)
    )
    """)
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS energeticos_empresa (
        id SERIAL PRIMARY KEY,
        empresa TEXT,
        energetico TEXT,
        proveedor TEXT,
        documento TEXT,
        FOREIGN KEY (empresa) REFERENCES usuarios(empresa)
    )
    """)
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS factores (
        categoria TEXT,
        unidad TEXT,
        factor REAL,
        PRIMARY KEY (categoria, unidad)
    )
    """)
    cursor.execute("CREATE TABLE IF NOT EXISTS agua_consumo (id SERIAL PRIMARY KEY, empresa TEXT, fecha TEXT, agua_embotellada_litros REAL, hielo_comprado_kg REAL, hielo_producido_kg REAL, tiene_tratamiento INTEGER DEFAULT 0, descripcion_tratamiento TEXT, FOREIGN KEY (empresa) REFERENCES usuarios(empresa))")
    cursor.execute("CREATE TABLE IF NOT EXISTS agua_afluentes (id SERIAL PRIMARY KEY, empresa TEXT, fecha TEXT, tipo TEXT, caudal_m3 REAL, tratamiento TEXT, FOREIGN KEY (empresa) REFERENCES usuarios(empresa))")
    cursor.execute("CREATE TABLE IF NOT EXISTS agua_cuencas (id SERIAL PRIMARY KEY, empresa TEXT, fecha TEXT, tipo_cuenca TEXT, cantidad_m3 REAL, FOREIGN KEY (empresa) REFERENCES usuarios(empresa))")
    cursor.execute("CREATE TABLE IF NOT EXISTS residuos_registros (id SERIAL PRIMARY KEY, empresa TEXT, fecha TEXT, periodo TEXT, tipo_residuo TEXT, cantidad_ton REAL, tratamiento TEXT, costo REAL, destino TEXT, FOREIGN KEY (empresa) REFERENCES usuarios(empresa))")
    cursor.execute("CREATE TABLE IF NOT EXISTS vehiculos (id SERIAL PRIMARY KEY, empresa TEXT, patente TEXT NOT NULL, tipo TEXT, marca TEXT, modelo TEXT, anio INTEGER, estado INTEGER DEFAULT 1, fecha_registro TEXT, FOREIGN KEY (empresa) REFERENCES usuarios(empresa))")
    cursor.execute("CREATE TABLE IF NOT EXISTS combustible_movil (id SERIAL PRIMARY KEY, empresa TEXT, vehiculo_id INTEGER, periodo TEXT, combustible TEXT, cantidad REAL, unidad TEXT, costo REAL, fecha_registro TEXT, FOREIGN KEY (empresa) REFERENCES usuarios(empresa))")
    cursor.execute("CREATE TABLE IF NOT EXISTS agua_costos (id SERIAL PRIMARY KEY, empresa TEXT, fecha TEXT, concepto TEXT, cantidad REAL, unidad TEXT, costo_usd REAL, costo_clp REAL, FOREIGN KEY (empresa) REFERENCES usuarios(empresa))")
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS factores_electricos (
        anio INTEGER,
        mes INTEGER,
        sistema TEXT,
        factor_emision_avg REAL
    )
    ''')

    cursor.execute("ALTER TABLE registros ADD COLUMN IF NOT EXISTS emision_ubicacion REAL DEFAULT 0")
    cursor.execute("ALTER TABLE registros ADD COLUMN IF NOT EXISTS origen_energia TEXT")
    cursor.execute("ALTER TABLE registros ADD COLUMN IF NOT EXISTS tiene_irec TEXT DEFAULT 'No'")
    cursor.execute("ALTER TABLE registros ADD COLUMN IF NOT EXISTS sistema TEXT")

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS irec_certificados (
        id SERIAL PRIMARY KEY,
        empresa TEXT,
        fecha_consumo TEXT,
        filename TEXT,
        contenido BYTEA,
        fecha_subida TEXT
    )
    """)
    cursor.execute("ALTER TABLE factores ADD COLUMN IF NOT EXISTS nombre_chile TEXT")
    cursor.execute("ALTER TABLE factores ADD COLUMN IF NOT EXISTS tratamiento TEXT")
    cursor.execute("ALTER TABLE factores ADD COLUMN IF NOT EXISTS año INTEGER DEFAULT 0")
    cursor.execute("UPDATE factores SET nombre_chile = NULL WHERE nombre_chile IN ('nan', '', 'None')")
    cursor.execute("UPDATE factores SET tratamiento = NULL WHERE tratamiento IN ('nan', '', 'None')")
    # Migrar residuos DEFRA de unidad 'kg' a 'tonne' (factor es kg CO₂e/t, no por kg)
    cursor.execute("""
        UPDATE factores SET unidad = 'tonne'
        WHERE unidad = 'kg' AND (nombre_chile IS NOT NULL OR tratamiento IS NOT NULL)
        AND (categoria, unidad, año) NOT IN (SELECT categoria, 'tonne', año FROM factores WHERE unidad = 'tonne')
    """)
    # Migrar PK a (categoria, unidad, año) y luego extender para incluir tratamiento
    try:
        cursor.execute("SAVEPOINT sp_factores_pk")
        cursor.execute("ALTER TABLE factores DROP CONSTRAINT factores_pkey")
        cursor.execute("RELEASE SAVEPOINT sp_factores_pk")
    except Exception:
        cursor.execute("ROLLBACK TO SAVEPOINT sp_factores_pk")
    # Índice único con tratamiento: permite un factor por (categoria, unidad, año, tratamiento)
    try:
        cursor.execute("SAVEPOINT sp_factores_trat_idx")
        cursor.execute("""
            CREATE UNIQUE INDEX IF NOT EXISTS factores_unique_trat
            ON factores (categoria, unidad, año, COALESCE(tratamiento, ''))
        """)
        cursor.execute("RELEASE SAVEPOINT sp_factores_trat_idx")
    except Exception:
        cursor.execute("ROLLBACK TO SAVEPOINT sp_factores_trat_idx")

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS pdf_uploads (
        id SERIAL PRIMARY KEY,
        empresa TEXT,
        fecha_subida TEXT,
        nombre_archivo TEXT,
        tipo TEXT,
        registros_generados INTEGER DEFAULT 0,
        sin_factor INTEGER DEFAULT 0
    )
    """)
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS pending_pdf_uploads (
        id SERIAL PRIMARY KEY,
        empresa TEXT,
        fecha_subida TEXT,
        nombre_archivo TEXT,
        tipo TEXT,
        datos_json TEXT,
        estado TEXT DEFAULT 'pendiente',
        motivo_rechazo TEXT,
        fecha_revision TEXT
    )
    """)

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS tickets (
        id SERIAL PRIMARY KEY,
        empresa TEXT NOT NULL,
        asunto TEXT NOT NULL,
        descripcion TEXT NOT NULL,
        prioridad TEXT DEFAULT 'Normal',
        estado TEXT DEFAULT 'Abierto',
        respuesta TEXT,
        fecha_creacion TEXT NOT NULL,
        fecha_respuesta TEXT
    )
    """)

    # Normalizar alcance existente basado en fuente (cubre NULL y valores incorrectos)
    cursor.execute("""
        UPDATE registros SET alcance = 'Alcance 1'
        WHERE fuente IN ('Combustión Fija','Combustible Móvil','Combustión Estacionaria','Refrigerantes','Fugas de Refrigerantes')
          AND (alcance IS NULL OR alcance NOT IN ('Alcance 1','Alcance 2','Alcance 3'))
    """)
    cursor.execute("""
        UPDATE registros SET alcance = 'Alcance 2'
        WHERE fuente = 'Electricidad'
          AND (alcance IS NULL OR alcance NOT IN ('Alcance 1','Alcance 2','Alcance 3'))
    """)
    cursor.execute("""
        UPDATE registros SET alcance = 'Alcance 3'
        WHERE fuente = 'Residuos'
          AND (alcance IS NULL OR alcance NOT IN ('Alcance 1','Alcance 2','Alcance 3'))
    """)

    admin_password = hashlib.sha256("admin123".encode()).hexdigest()
    cursor.execute("""
        INSERT INTO usuarios
        (empresa, email, password, contacto, es_admin, fecha_registro)
        VALUES (%s, %s, %s, %s, %s, %s)
        ON CONFLICT (email) DO UPDATE SET password = EXCLUDED.password
    """, ("Administrador", "admin@huella.com", admin_password, "Administrador", 1, datetime.now().strftime("%Y-%m-%d %H:%M")))

    conn.commit()
    conn.close()

# ================= DASHBOARD PRINCIPAL =================
@app.route("/dashboard")
def dashboard():
    if 'user_id' not in session or session.get('es_admin') == 1: return redirect("/")
    empresa = session.get('empresa')
    conn = get_db()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    
    # 1. BUSCADOR DE AÑOS
    cursor.execute("""
        SELECT DISTINCT SUBSTRING(fecha::text, 1, 4) as anio FROM registros 
        WHERE empresa = %s AND fecha IS NOT NULL AND length(fecha::text) >= 4
        UNION
        SELECT DISTINCT SUBSTRING(fecha_registro::text, 1, 4) FROM combustible_movil 
        WHERE empresa = %s AND fecha_registro IS NOT NULL AND length(fecha_registro::text) >= 4
        ORDER BY anio DESC
    """, (empresa, empresa))
    anios_disponibles = [row['anio'] for row in cursor.fetchall() if row['anio']]

    cursor.execute("SELECT anio_default FROM usuarios WHERE id = %s", (session['user_id'],))
    pref = cursor.fetchone()
    anio_default = pref['anio_default'] if pref and pref['anio_default'] else 'Todos'
    anio_seleccionado = request.args.get('anio', anio_default)
    params_reg = [empresa]
    params_comb = [empresa]
    filtro_fecha_reg = ""
    filtro_fecha_comb = ""

    # FILTRO: Usamos cm.fecha_registro para evitar ambigüedades en el JOIN
    if anio_seleccionado != 'Todos':
        filtro_fecha_reg = " AND SUBSTRING(fecha::text, 1, 4) = %s"
        filtro_fecha_comb = " AND SUBSTRING(cm.fecha_registro::text, 1, 4) = %s"
        params_reg.append(anio_seleccionado)
        params_comb.append(anio_seleccionado)

    # 2. TOTALES Y VEHÍCULOS
    cursor.execute(f"SELECT SUM(emision) FROM registros WHERE empresa = %s {filtro_fecha_reg}", params_reg)
    res_reg = cursor.fetchone()
    total_reg = float(res_reg[0]) if res_reg and res_reg[0] else 0.0
    total_comb = 0.0  # incluido en registros desde ahora

    # Breakdown por vehículo (para la tarjeta de flota en el dashboard)
    cursor.execute(f"SELECT cm.cantidad, cm.combustible, v.patente, v.tipo FROM combustible_movil cm LEFT JOIN vehiculos v ON cm.vehiculo_id = v.id WHERE cm.empresa = %s {filtro_fecha_comb}", params_comb)
    comb_data = cursor.fetchall()
    vehiculos_dict = {}
    factores_v = {"diesel": 2.68, "bencina": 2.31, "glp": 1.61, "gas_natural": 2.02, "electricidad": 0.233}
    for row in comb_data:
        f = factores_v.get(str(row['combustible']).lower(), 2.5)
        emision = (float(row['cantidad']) or 0.0) * f
        pat = row['patente'] or 'Sin Asignar'
        if pat not in vehiculos_dict:
            vehiculos_dict[pat] = {'patente': pat, 'tipo': row['tipo'] or 'N/A', 'registros': 0, 'total_emision': 0.0}
        vehiculos_dict[pat]['registros'] += 1
        vehiculos_dict[pat]['total_emision'] += emision

    cursor.execute("SELECT * FROM vehiculos WHERE empresa = %s", (empresa,))
    vehiculos = cursor.fetchall()

    # 3. GRÁFICO (Conversión a Float para no romper el JSON)
    cursor.execute(f"SELECT fuente, SUM(emision) as total FROM registros WHERE empresa = %s {filtro_fecha_reg} GROUP BY fuente", params_reg)
    categorias = [{'fuente': row['fuente'], 'total': float(row['total']) if row['total'] else 0.0} for row in cursor.fetchall()]

    # 4. CONTEO TOTAL Y ÚLTIMOS REGISTROS
    cursor.execute(f"SELECT COUNT(*) FROM registros WHERE empresa = %s {filtro_fecha_reg}", params_reg)
    total_registros = cursor.fetchone()[0] or 0

    cursor.execute(f"""
        SELECT id, fecha, fuente, actividad, cantidad, unidad, emision, 'manual' as tipo_tabla
        FROM registros WHERE empresa = %s {filtro_fecha_reg}
        ORDER BY fecha DESC LIMIT 5
    """, params_reg)
    ultimos = [dict(row) for row in cursor.fetchall()]

    # 5. TENDENCIA MENSUAL POR ALCANCE (solo registros — combustible_movil ya está incluido)
    cursor.execute(f"""
        SELECT
            SUBSTRING(fecha::text, 1, 7) as mes,
            CASE
                WHEN fuente IN ('Combustión Estacionaria', 'Combustión Fija', 'Combustible Móvil', 'Combustión Móvil', 'Refrigerantes', 'Fugas de Refrigerantes') THEN 1
                WHEN fuente = 'Electricidad' THEN 2
                WHEN fuente = 'Residuos' THEN 3
                ELSE 1
            END as alcance,
            SUM(emision) as total
        FROM registros WHERE empresa = %s {filtro_fecha_reg}
        GROUP BY 1, 2 ORDER BY 1
    """, params_reg)
    tendencia_reg = cursor.fetchall()

    from collections import defaultdict
    meses_dict = defaultdict(lambda: {1: 0.0, 2: 0.0, 3: 0.0})
    for row in tendencia_reg:
        meses_dict[row['mes']][row['alcance']] += float(row['total'] or 0)

    meses_sorted = sorted(meses_dict.keys())
    tendencia_labels = meses_sorted
    tendencia_a1 = [round(meses_dict[m][1], 2) for m in meses_sorted]
    tendencia_a2 = [round(meses_dict[m][2], 2) for m in meses_sorted]
    tendencia_a3 = [round(meses_dict[m][3], 2) for m in meses_sorted]

    alcance_a1 = round(sum(tendencia_a1), 2)
    alcance_a2 = round(sum(tendencia_a2), 2)
    alcance_a3 = round(sum(tendencia_a3), 2)
    alcances_data = [
        {'nombre': 'Alcance 1 – Directo', 'total': alcance_a1, 'color': '#EF4444'},
        {'nombre': 'Alcance 2 – Electricidad', 'total': alcance_a2, 'color': '#F59E0B'},
        {'nombre': 'Alcance 3 – Residuos', 'total': alcance_a3, 'color': '#10B981'},
    ]

    # 6. VARIACIÓN INTERANUAL
    variacion_pct = None
    anio_prev = None
    if anio_seleccionado != 'Todos' and anio_seleccionado:
        anio_prev = str(int(anio_seleccionado) - 1)
        cursor.execute("SELECT SUM(emision) FROM registros WHERE empresa = %s AND SUBSTRING(fecha::text, 1, 4) = %s", (empresa, anio_prev))
        res_p = cursor.fetchone()
        total_prev = float(res_p[0]) if res_p and res_p[0] else 0.0
        total_actual = total_reg
        if total_prev > 0:
            variacion_pct = round((total_actual - total_prev) / total_prev * 100, 1)

    conn.close()

    return render_template("dashboard.html",
                           total_emision=total_reg + total_comb, total_emision_registros=total_reg, total_emision_combustible=total_comb,
                           categorias_data=categorias, ultimos_registros=ultimos, total_registros=total_registros, empresa=empresa,
                           anios_disponibles=anios_disponibles, anio_seleccionado=anio_seleccionado,
                           vehiculos_emisiones=list(vehiculos_dict.values()), tiene_vehiculos=len(vehiculos)>0, vehiculos=vehiculos,
                           tendencia_labels=tendencia_labels, tendencia_a1=tendencia_a1,
                           tendencia_a2=tendencia_a2, tendencia_a3=tendencia_a3,
                           alcances_data=alcances_data, variacion_pct=variacion_pct, anio_prev=anio_prev)

# ================= HISTORIAL COMPLETO =================
@app.route("/historial")
def historial():
    if 'user_id' not in session: return redirect("/")
    empresa = session.get('empresa')
    POR_PAGINA = 50
    pagina = max(1, int(request.args.get('pagina', 1)))

    filtro_alcance = request.args.get('alcance', '').strip()
    filtro_fuente  = request.args.get('fuente', '').strip()
    filtro_anio    = request.args.get('anio', '').strip()
    filtro_q       = request.args.get('q', '').strip()

    conn = get_db()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    # Opciones para los filtros desplegables
    alcances_disponibles = ['Alcance 1', 'Alcance 2', 'Alcance 3']
    cursor.execute("SELECT DISTINCT fuente FROM registros WHERE empresa = %s AND fuente IS NOT NULL ORDER BY fuente", (empresa,))
    fuentes_disponibles = [r[0] for r in cursor.fetchall()]
    cursor.execute("SELECT DISTINCT SUBSTRING(fecha::text,1,4) as anio FROM registros WHERE empresa = %s AND fecha IS NOT NULL ORDER BY anio DESC", (empresa,))
    anios_disponibles = [r[0] for r in cursor.fetchall()]

    _fuentes_a1 = ('Combustión Fija', 'Combustible Móvil', 'Combustión Estacionaria', 'Refrigerantes', 'Fugas de Refrigerantes')
    _fuentes_a2 = ('Electricidad',)
    _fuentes_a3 = ('Residuos',)

    # WHERE dinámico
    where = ["empresa = %s"]
    params = [empresa]
    if filtro_alcance == 'Alcance 1':
        where.append("fuente IN %s"); params.append(_fuentes_a1)
    elif filtro_alcance == 'Alcance 2':
        where.append("fuente IN %s"); params.append(_fuentes_a2)
    elif filtro_alcance == 'Alcance 3':
        where.append("fuente IN %s"); params.append(_fuentes_a3)
    if filtro_fuente:
        where.append("fuente = %s"); params.append(filtro_fuente)
    if filtro_anio:
        where.append("SUBSTRING(fecha::text,1,4) = %s"); params.append(filtro_anio)
    if filtro_q:
        where.append("(LOWER(fuente) LIKE %s OR LOWER(actividad) LIKE %s OR LOWER(categoria) LIKE %s)")
        like = f"%{filtro_q.lower()}%"
        params += [like, like, like]

    where_sql = " AND ".join(where)

    cursor.execute(f"SELECT COUNT(*) FROM registros WHERE {where_sql}", params)
    total = cursor.fetchone()[0]
    total_paginas = max(1, (total + POR_PAGINA - 1) // POR_PAGINA)
    pagina = min(pagina, total_paginas)
    offset = (pagina - 1) * POR_PAGINA

    cursor.execute(f"""
        SELECT id, fecha, alcance, fuente, categoria, actividad, cantidad, unidad, emision, 'manual' as tipo_tabla
        FROM registros WHERE {where_sql}
        ORDER BY fecha DESC, id DESC
        LIMIT %s OFFSET %s
    """, params + [POR_PAGINA, offset])

    registros = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return render_template("historial.html", registros=registros, empresa=empresa,
                           pagina=pagina, total_paginas=total_paginas, total=total,
                           alcances_disponibles=alcances_disponibles,
                           fuentes_disponibles=fuentes_disponibles,
                           anios_disponibles=anios_disponibles,
                           filtro_alcance=filtro_alcance, filtro_fuente=filtro_fuente,
                           filtro_anio=filtro_anio, filtro_q=filtro_q)


# ================= PANELES INDIVIDUALES =================
@app.route("/combustion")
def combustion_dashboard():
    if 'user_id' not in session:
        return redirect("/")
    
    empresa = session.get('empresa')
    conn = get_db()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    
    fuentes_comb = ('Combustión Estacionaria', 'Combustión Fija', 'Combustión Móvil', 'Combustible Móvil')
    
    cursor.execute("SELECT fecha, fuente, categoria, actividad, cantidad, unidad, emision FROM registros WHERE empresa = %s AND fuente IN %s ORDER BY fecha DESC LIMIT 10", (empresa, fuentes_comb))
    registros_comb = [dict(row) for row in cursor.fetchall()]
    
    cursor.execute("SELECT categoria, SUM(emision) as total FROM registros WHERE empresa = %s AND fuente IN %s GROUP BY categoria", (empresa, fuentes_comb))
    grafico_data = [dict(row) for row in cursor.fetchall()]

    cursor.execute("SELECT SUM(emision) FROM registros WHERE empresa = %s AND fuente IN %s", (empresa, fuentes_comb))
    res = cursor.fetchone()
    total_emision = res[0] if res and res[0] else 0

    cursor.execute("""
        SELECT SUBSTRING(fecha::text, 1, 7) as mes, fuente, SUM(emision) as total
        FROM registros WHERE empresa = %s AND fuente IN %s
        GROUP BY 1, 2 ORDER BY 1
    """, (empresa, fuentes_comb))
    tendencia_comb_raw = cursor.fetchall()
    meses_comb = sorted({row['mes'] for row in tendencia_comb_raw})
    tendencia_fija = []
    tendencia_movil = []
    for m in meses_comb:
        fija = sum(float(r['total'] or 0) for r in tendencia_comb_raw if r['mes'] == m and 'Fija' in r['fuente'])
        movil = sum(float(r['total'] or 0) for r in tendencia_comb_raw if r['mes'] == m and 'Fija' not in r['fuente'])
        tendencia_fija.append(round(fija, 2))
        tendencia_movil.append(round(movil, 2))

    conn.close()
    return render_template("combustion_dashboard.html", registros=registros_comb, grafico_data=grafico_data,
                           total_emision=total_emision, meses_comb=meses_comb,
                           tendencia_fija=tendencia_fija, tendencia_movil=tendencia_movil)

@app.route("/electricidad")
def electricidad_dashboard():
    if 'user_id' not in session: return redirect("/")

    empresa = session.get('empresa')
    conn = get_db()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    cursor.execute(
        "SELECT fecha, categoria, actividad, cantidad, unidad, emision, emision_ubicacion, "
        "origen_energia, tiene_irec, COALESCE(sistema, '') as sistema, factor "
        "FROM registros WHERE empresa = %s AND fuente = 'Electricidad' ORDER BY fecha DESC LIMIT 20",
        (empresa,)
    )
    rows = cursor.fetchall()
    registros_elec = []
    for r in rows:
        d = dict(r)
        cant = d['cantidad'] or 0
        d['factor_ubicacion'] = round(d['emision_ubicacion'] / cant, 6) if cant else 0.0
        registros_elec.append(d)
    cursor.execute(
        "SELECT sistema, SUM(emision) as total FROM registros WHERE empresa = %s AND fuente = 'Electricidad' GROUP BY sistema",
        (empresa,)
    )
    grafico_data_raw = cursor.fetchall()
    if not grafico_data_raw:
        cursor.execute(
            "SELECT categoria, SUM(emision) as total FROM registros WHERE empresa = %s AND fuente = 'Electricidad' GROUP BY categoria",
            (empresa,)
        )
        grafico_data_raw = cursor.fetchall()
    grafico_data = [dict(r) for r in grafico_data_raw]

    cursor.execute("SELECT SUM(emision) as total FROM registros WHERE empresa = %s AND fuente = 'Electricidad'", (empresa,))
    res = cursor.fetchone()
    total_mercado = float(res['total']) if res and res['total'] else 0.0

    cursor.execute("SELECT SUM(emision_ubicacion) as total FROM registros WHERE empresa = %s AND fuente = 'Electricidad'", (empresa,))
    res2 = cursor.fetchone()
    total_ubicacion = float(res2['total']) if res2 and res2['total'] else 0.0

    cursor.execute("SELECT SUM(cantidad) as total FROM registros WHERE empresa = %s AND fuente = 'Electricidad'", (empresa,))
    res3 = cursor.fetchone()
    total_kwh = float(res3['total']) if res3 and res3['total'] else 0.0

    cursor.execute("SELECT COUNT(*) as total FROM registros WHERE empresa = %s AND fuente = 'Electricidad' AND tiene_irec = 'Si'", (empresa,))
    total_irec = cursor.fetchone()['total'] or 0

    cursor.execute("""
        SELECT SUBSTRING(fecha::text, 1, 7) as mes,
               SUM(cantidad) as kwh,
               SUM(emision) as emision_mercado,
               SUM(emision_ubicacion) as emision_ubicacion
        FROM registros WHERE empresa = %s AND fuente = 'Electricidad'
        GROUP BY 1 ORDER BY 1
    """, (empresa,))
    tendencia_elec = [dict(r) for r in cursor.fetchall()]

    conn.close()
    return render_template(
        "electricidad_dashboard.html",
        registros=registros_elec,
        grafico_data=grafico_data,
        total_emision=total_mercado,
        total_mercado=total_mercado,
        total_ubicacion=total_ubicacion,
        total_kwh=total_kwh,
        total_irec=total_irec,
        tendencia_elec=tendencia_elec,
    )

@app.route("/agua")
def agua_dashboard():
    if 'user_id' not in session: return redirect("/")
    
    empresa = session.get('empresa')
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute("SELECT fecha, agua_embotellada_litros, hielo_comprado_kg, hielo_producido_kg FROM agua_consumo WHERE empresa = %s ORDER BY fecha DESC LIMIT 5", (empresa,))
    consumo_data = cursor.fetchall()
    
    cursor.execute("SELECT tipo_cuenca, SUM(cantidad_m3) FROM agua_cuencas WHERE empresa = %s GROUP BY tipo_cuenca", (empresa,))
    cuencas_data = cursor.fetchall()
    
    cursor.execute("SELECT SUM(agua_embotellada_litros * 0.25 + hielo_comprado_kg * 0.18 + hielo_producido_kg * 0.15) FROM agua_consumo WHERE empresa = %s", (empresa,))
    res = cursor.fetchone()
    total_emision = res[0] if res and res[0] else 0
    
    cursor.execute("SELECT SUM(costo_usd), SUM(costo_clp) FROM agua_costos WHERE empresa = %s", (empresa,))
    costos = cursor.fetchone()
    total_usd = costos[0] if costos and costos[0] else 0
    total_clp = costos[1] if costos and costos[1] else 0
    
    conn.close()
    return render_template("agua_dashboard.html", empresa=empresa, consumo_data=consumo_data, cuencas_data=cuencas_data, total_emision=total_emision, total_usd=total_usd, total_clp=total_clp)

@app.route("/refrigerantes")
def refrigerantes_dashboard():
    if 'user_id' not in session: return redirect("/")
    
    empresa = session.get('empresa')
    conn = get_db()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    
    fuentes_ref = ('Refrigerantes', 'Fugas de Refrigerantes')
    cursor.execute("SELECT SUM(emision) as total FROM registros WHERE empresa = %s AND fuente IN %s", (empresa, fuentes_ref))
    res = cursor.fetchone()
    total_refrigerantes = res['total'] if res and res['total'] else 0.0

    cursor.execute("SELECT categoria, SUM(emision) as total FROM registros WHERE empresa = %s AND fuente IN %s GROUP BY categoria ORDER BY total DESC", (empresa, fuentes_ref))
    datos_gases = [dict(row) for row in cursor.fetchall()]

    cursor.execute("SELECT SUBSTRING(fecha, 1, 7) as mes, SUM(emision) as total FROM registros WHERE empresa = %s AND fuente IN %s GROUP BY mes ORDER BY mes", (empresa, fuentes_ref))
    datos_meses = [dict(row) for row in cursor.fetchall()]

    cursor.execute("SELECT fecha, categoria, cantidad, unidad, factor, emision FROM registros WHERE empresa = %s AND fuente IN %s ORDER BY fecha DESC LIMIT 10", (empresa, fuentes_ref))
    rows = cursor.fetchall()

    # Cargar factores de la tabla factores para recuperar emisiones de registros con factor=0
    cursor.execute("SELECT categoria, unidad, factor FROM factores WHERE factor > 0")
    factores_lookup = {}
    for frow in cursor.fetchall():
        factores_lookup[(frow['categoria'], frow['unidad'])] = float(frow['factor'])

    ultimos_registros = []
    for row in rows:
        r = dict(row)
        cant = r.get('cantidad') or 0
        em = r.get('emision') or 0
        fac = r.get('factor') or 0

        if cant == 0 and em > 0 and fac > 0:
            # Caso 1: cantidad perdida, recuperar desde emision/factor
            r['cantidad'] = round(em / fac, 6)
        elif em == 0 and cant > 0:
            # Caso 2: emision guardada como 0 (factor era 0 al guardar), buscar factor real
            fac_real = (factores_lookup.get((r['categoria'], r['unidad'])) or
                        factores_lookup.get((r['categoria'], 'kg')))
            if fac_real:
                r['emision'] = round(cant * fac_real, 4)
                r['factor'] = fac_real
        ultimos_registros.append(r)

    cursor.execute("SELECT COUNT(DISTINCT categoria) as total FROM registros WHERE empresa = %s AND fuente IN %s", (empresa, fuentes_ref))
    res_gases = cursor.fetchone()
    total_gases_distintos = res_gases['total'] if res_gases and res_gases['total'] else 0

    gas_top = datos_gases[0] if datos_gases else None

    conn.close()
    return render_template("refrigerantes_dashboard.html", total_refrigerantes=total_refrigerantes,
                           datos_gases=datos_gases, datos_meses=datos_meses, registros=ultimos_registros,
                           total_gases_distintos=total_gases_distintos, gas_top=gas_top)


# ================= GESTIÓN DE RESIDUOS (EL EXTRACTOR MÁGICO) =================
@app.route('/residuos', methods=['GET'])
def residuos_dashboard():
    if 'user_id' not in session: return redirect("/")
    empresa = session.get('empresa')
    conn = get_db()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    
    cursor.execute("SELECT SUM(emision) as total FROM registros WHERE empresa = %s AND fuente = 'Residuos'", (empresa,))
    res = cursor.fetchone()
    total = float(res['total']) if res and res['total'] else 0.0
    
    cursor.execute("SELECT categoria, SUM(emision) as total FROM registros WHERE empresa = %s AND fuente = 'Residuos' GROUP BY categoria", (empresa,))
    datos_cat = [{'categoria': row['categoria'], 'total': float(row['total']) if row['total'] else 0.0} for row in cursor.fetchall()]
    
    cursor.execute("SELECT SUBSTRING(fecha, 1, 7) as mes, SUM(emision) as total FROM registros WHERE empresa = %s AND fuente = 'Residuos' GROUP BY mes ORDER BY mes", (empresa,))
    datos_mes = [{'mes': row['mes'], 'total': float(row['total']) if row['total'] else 0.0} for row in cursor.fetchall()]

    cursor.execute("""
        SELECT COALESCE(NULLIF(TRIM(actividad), ''), 'Sin especificar') as trat,
               SUM(emision) as total
        FROM registros WHERE empresa = %s AND fuente = 'Residuos'
        GROUP BY 1 ORDER BY 2 DESC
    """, (empresa,))
    datos_tratamiento = [{'tratamiento': row['trat'], 'total': float(row['total']) if row['total'] else 0.0} for row in cursor.fetchall()]

    conn.close()
    return render_template("residuos.html", total_emision=total, datos_categoria=datos_cat,
                           datos_mes=datos_mes, datos_tratamiento=datos_tratamiento)

# ================= FORMULARIO RESIDUOS (CON BUSCADOR INTELIGENTE) =================
@app.route("/formulario_residuos", methods=['GET', 'POST'])
def formulario_residuos():
    import re
    import pandas as pd
    import unicodedata
    from datetime import datetime

    if 'user_id' not in session: return redirect("/")

    def limpiar_texto(texto):
        if not texto: return ""
        texto = str(texto).lower().strip()
        return ''.join(c for c in unicodedata.normalize('NFD', texto) if unicodedata.category(c) != 'Mn')

    def sin_tildes(texto):
        return ''.join(c for c in unicodedata.normalize('NFD', str(texto).lower()) if unicodedata.category(c) != 'Mn')

    _TRAT_SINONIMOS = [
        (['monorelleno', 'relleno sanitario', 'disposicion final', 'vertedero'], 'Vertedero'),
        (['reciclaje', 'reciclado', 'open-loop', 'open loop'], 'Reciclaje'),
        (['compostaje', 'compost'], 'Compostaje'),
        (['coincineracion', 'coincineración', 'combustion', 'combustión', 'incineracion', 'incineración'], 'Combustión'),
        (['digestion anaerobica', 'digestión anaeróbica', 'anaerobica', 'anaeróbica'], 'Digestión anaeróbica'),
        (['pretratamiento'], 'Pretratamiento'),
        (['aplicacion a suelo', 'aplicación a suelo'], 'Aplicación a suelo'),
    ]

    def normalizar_tratamiento(texto):
        t = limpiar_texto(texto)
        if not t or t in ('nan', 'none'):
            return ''
        for palabras_clave, nombre_normalizado in _TRAT_SINONIMOS:
            if any(p in t for p in palabras_clave):
                return nombre_normalizado
        return texto.strip()

    def get_factores_residuos():
        conn2 = get_db()
        cur2 = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
        cur2.execute("""
            SELECT categoria, unidad, factor, tratamiento,
                   CASE WHEN TRIM(LOWER(COALESCE(nombre_chile,''))) IN ('nan','none','') THEN NULL ELSE nombre_chile END as nombre_chile,
                   COALESCE(año, 0) as año
            FROM factores ORDER BY año DESC, categoria, tratamiento
        """)
        todos = [dict(r) for r in cur2.fetchall()]
        conn2.close()
        palabras_clave = ['residuo', 'papel', 'carton', 'plastico', 'vidrio', 'metal', 'organico',
                          'wood', 'glass', 'batteries', 'weee', 'clothing', 'tyres', 'mineral', 'organic']
        filtrados = [f for f in todos if any(p in sin_tildes(f['categoria']) for p in palabras_clave)]
        # Deduplicate by (categoria, tratamiento): keep the latest año per combination
        seen = set()
        unique = []
        for f in filtrados:
            key = (f['categoria'], f.get('tratamiento') or '')
            if key not in seen:
                seen.add(key)
                unique.append(f)
        return unique

    def buscar_factor(residuo_limpio, factores_db, año=0, tratamiento=None):
        residuo_norm = limpiar_texto(residuo_limpio)
        defra_cat = clasificar_defra(residuo_limpio)
        defra_norm = limpiar_texto(defra_cat)
        trat_norm = limpiar_texto(tratamiento or '')
        def _hit(f):
            return float(f['factor']), defra_cat, f.get('nombre_chile', '') or '', f.get('tratamiento', '') or ''
        stopwords = {'de', 'y', 'en', 'para', 'los', 'las', 'el', 'la', 'con', 'sin', 'tipo',
                     'residuos', 'residuo', 'envases', 'mezcla', 'mezclas', 'otros', 'materiales'}
        palabras_pdf = set(w for w in residuo_norm.replace(',', ' ').replace('.', ' ').split()
                           if w not in stopwords and len(w) > 2)
        def _trat_match(f_trat):
            if not trat_norm or not f_trat:
                return False
            f_trat_n = limpiar_texto(f_trat)
            return trat_norm in f_trat_n or f_trat_n in trat_norm
        def _search(db, require_trat=False):
            # Primera pasada: match exacto por categoría DEFRA (con tratamiento si aplica)
            for f in db:
                if limpiar_texto(str(f['categoria'])) == defra_norm:
                    if not require_trat or _trat_match(f.get('tratamiento', '')):
                        return _hit(f)
            # Segunda pasada: match parcial por nombre de categoría
            for f in db:
                cat_db_norm = limpiar_texto(f['categoria'])
                coincide_cat = (cat_db_norm == residuo_norm or cat_db_norm in residuo_norm
                                or residuo_norm in cat_db_norm)
                if not coincide_cat:
                    palabras_db = set(w for w in cat_db_norm.replace(',', ' ').replace('.', ' ').split()
                                      if w not in stopwords and len(w) > 2)
                    coincide_cat = bool(palabras_db and palabras_db.intersection(palabras_pdf))
                if coincide_cat:
                    if not require_trat or _trat_match(f.get('tratamiento', '')):
                        return _hit(f)
            return None
        # Buscar primero con año y tratamiento, luego sin tratamiento, luego sin año
        for pool in (
            [f for f in factores_db if (f.get('año') or 0) == año] if año else [],
            factores_db,
        ):
            if not pool:
                continue
            if trat_norm:
                result = _search(pool, require_trat=True)
                if result:
                    return result
            result = _search(pool, require_trat=False)
            if result:
                return result
        return (0.0, defra_cat, '', '')

    if request.method == "POST":
        tipo_ingreso = request.form.get("tipo_ingreso", "manual")
        empresa = session.get('empresa')

        # === PASO 2: enviar a revisión del admin ===
        if tipo_ingreso == 'confirmar_pdf':
            datos_json = request.form.get('datos_confirmados', '[]')
            nombre_pdf = request.form.get('nombre_pdf', 'desconocido')
            tipo_pdf = request.form.get('tipo_pdf', 'pdf')
            periodo_override = request.form.get('periodo_override', '').strip()  # YYYY-MM
            if periodo_override:
                try:
                    year, month = periodo_override.split('-')
                    fecha_override = f"{year}-{month}-01"
                    filas = json.loads(datos_json)
                    for f in filas:
                        f['fecha'] = fecha_override
                    datos_json = json.dumps(filas)
                except Exception:
                    pass
            try:
                conn = get_db()
                cursor = conn.cursor()
                cursor.execute("""
                    INSERT INTO pending_pdf_uploads (empresa, fecha_subida, nombre_archivo, tipo, datos_json, estado)
                    VALUES (%s, %s, %s, %s, %s, 'pendiente')
                """, (empresa, datetime.now().strftime("%Y-%m-%d %H:%M"), nombre_pdf, tipo_pdf, datos_json))
                conn.commit()
                conn.close()
                flash("Tu archivo fue enviado para revisión. El administrador lo validará y recibirás confirmación pronto.", "info")
            except Exception as e:
                flash(f"Error al enviar: {str(e)}", "danger")
            return redirect("/residuos")

        # === PASO 1: procesar PDF y mostrar preview ===
        if tipo_ingreso in ['sinader', 'sidrep']:
            archivo_pdf = request.files.get("archivo_pdf")
            if not archivo_pdf or archivo_pdf.filename == '':
                flash("Debes subir un archivo PDF válido.", "danger")
                return redirect(request.url)
            try:
                temp_path = os.path.join(tempfile.gettempdir(), archivo_pdf.filename)
                archivo_pdf.save(temp_path)
                if tipo_ingreso == 'sinader':
                    df_extraido = extract_sinader_data(temp_path)
                    col_res, col_trat, col_cant, col_dest = 'Residuo', 'Tipo Tratamiento', 'Cantidad (kg)', 'Destino'
                else:
                    df_extraido = extract_sidrep_data(temp_path)
                    col_res, col_trat, col_cant, col_dest = 'Descripción Residuo', 'Estado del Residuo', 'Cantidad (Kg)', 'Empresa destinataria'
                os.remove(temp_path)

                conn = get_db()
                cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
                cursor.execute("SELECT categoria, factor, nombre_chile, tratamiento, COALESCE(año, 0) as año FROM factores")
                factores_db = cursor.fetchall()
                conn.close()

                # Extraer año y mes del PDF para asignar fecha correcta a los registros
                año_pdf = datetime.now().year
                mes_pdf = None  # None = desconocido
                periodo_detectado = ""  # YYYY-MM para el input type=month
                try:
                    if 'Periodo' in df_extraido.columns and len(df_extraido) > 0:
                        # SINADER: Periodo tiene formato MM/YYYY
                        periodo_str = str(df_extraido['Periodo'].iloc[0])
                        m_anio = re.search(r'\b(20\d{2})\b', periodo_str)
                        m_mes = re.match(r'^(\d{2})/', periodo_str.strip())
                        if m_anio:
                            año_pdf = int(m_anio.group(1))
                        if m_mes:
                            mes_pdf = int(m_mes.group(1))
                    elif 'FechaDeclaración' in df_extraido.columns and len(df_extraido) > 0:
                        # SIDREP: FechaDeclaración tiene formato YYYY-MM-DD
                        fecha_str = str(df_extraido['FechaDeclaración'].iloc[0])
                        m_fecha = re.match(r'^(\d{4})-(\d{2})', fecha_str)
                        if m_fecha:
                            año_pdf = int(m_fecha.group(1))
                            mes_pdf = int(m_fecha.group(2))
                except Exception:
                    pass

                if mes_pdf:
                    fecha_registro_pdf = f"{año_pdf}-{mes_pdf:02d}-01"
                    periodo_detectado = f"{año_pdf}-{mes_pdf:02d}"
                else:
                    fecha_registro_pdf = f"{año_pdf}-01-01"
                    periodo_detectado = f"{año_pdf}-01"

                preview_rows = []
                for _, row in df_extraido.iterrows():
                    residuo_raw = str(row.get(col_res, 'Desconocido'))
                    destino = str(row.get(col_dest, ''))
                    residuo_limpio = re.sub(r'^[\d\s]+', '', residuo_raw).strip()

                    raw_cant = row.get(col_cant, 0)
                    cantidad = 0.0
                    if pd.isna(raw_cant): cantidad = 0.0
                    elif isinstance(raw_cant, (int, float)): cantidad = float(raw_cant)
                    else:
                        txt = str(raw_cant).lower().replace('kg', '').strip()
                        if '.' in txt and ',' in txt: txt = txt.replace('.', '').replace(',', '.')
                        elif ',' in txt: txt = txt.replace(',', '.')
                        elif '.' in txt and len(txt.split('.')[-1]) == 3: txt = txt.replace('.', '')
                        try: cantidad = float(txt)
                        except: cantidad = 0.0

                    if cantidad <= 0: continue

                    if tipo_ingreso == 'sidrep':
                        trat_pdf = 'Vertedero'
                    else:
                        trat_pdf = normalizar_tratamiento(str(row.get(col_trat, '')))
                    factor, defra_cat, nombre_chile, tratamiento_defra = buscar_factor(
                        residuo_limpio, factores_db, año=año_pdf, tratamiento=trat_pdf)
                    if not trat_pdf:
                        trat_pdf = tratamiento_defra or ''
                    emision = round((cantidad / 1000) * factor, 4)
                    preview_rows.append({
                        'fecha': fecha_registro_pdf,
                        'destino': (destino or 'Operaciones')[:50],
                        'categoria': residuo_limpio,
                        'nombre_chile': nombre_chile,
                        'tratamiento': trat_pdf,
                        'cantidad': cantidad,
                        'factor': factor,
                        'emision': emision,
                        'defra_cat': defra_cat,
                    })

                conn2 = get_db()
                cur2 = conn2.cursor()
                cur2.execute("SELECT DISTINCT tratamiento FROM factores WHERE tratamiento IS NOT NULL AND tratamiento != '' ORDER BY tratamiento")
                tratamientos_disponibles = [r[0] for r in cur2.fetchall()]
                conn2.close()

                return render_template("formulario_residuos.html",
                                       factores=get_factores_residuos(),
                                       preview_rows=preview_rows,
                                       preview_json=json.dumps(preview_rows),
                                       tipo_ingreso_prev=tipo_ingreso,
                                       nombre_pdf=archivo_pdf.filename,
                                       periodo_detectado=periodo_detectado,
                                       tratamientos_disponibles=tratamientos_disponibles)
            except Exception as e:
                flash(f"Error procesando el PDF. Detalle: {str(e)}", "danger")
                return redirect(request.url)

        # === MANUAL ===
        conn = get_db()
        cursor = conn.cursor()
        periodos = request.form.getlist("periodo[]")
        tipos = request.form.getlist("tipo_residuo[]")
        cantidades = request.form.getlist("cantidad[]")
        tratamientos = request.form.getlist("tratamiento[]")
        factores_filas = request.form.getlist("factor[]")
        destinos = request.form.getlist("destino[]")

        # Preload factors: {(categoria, tratamiento_lower, año): factor}
        cursor.execute("SELECT categoria, COALESCE(tratamiento, '') as trat, COALESCE(año, 0) as año, factor FROM factores")
        _fac_rows = cursor.fetchall()
        factores_lookup = {}
        for _cat, _trat, _año, _fac in _fac_rows:
            factores_lookup[(_cat, _trat.lower(), _año)] = _fac
            if (_cat, _trat.lower(), 0) not in factores_lookup:
                factores_lookup[(_cat, _trat.lower(), 0)] = _fac
            # Fallback sin tratamiento
            if (_cat, '', _año) not in factores_lookup:
                factores_lookup[(_cat, '', _año)] = _fac
            if (_cat, '', 0) not in factores_lookup:
                factores_lookup[(_cat, '', 0)] = _fac

        filas_guardadas = 0
        for i in range(len(periodos)):
            if not periodos[i].strip() or not cantidades[i].strip(): continue
            fecha_limpia = f"{periodos[i]}-01"
            try:
                cant = float(cantidades[i].replace(',', '.'))
                año_periodo = int(periodos[i][:4]) if len(periodos[i]) >= 4 else datetime.now().year
                trat_raw = normalizar_tratamiento(tratamientos[i] if i < len(tratamientos) else '')
                trat_key = trat_raw.lower()
                fac = (factores_lookup.get((tipos[i], trat_key, año_periodo))
                       or factores_lookup.get((tipos[i], trat_key, 0))
                       or factores_lookup.get((tipos[i], '', año_periodo))
                       or factores_lookup.get((tipos[i], '', 0))
                       or float(factores_filas[i].replace(',', '.') if i < len(factores_filas) else 0))
            except: cant, fac = 0.0, 0.0
            emision = (cant / 1000) * fac
            destino = destinos[i] if i < len(destinos) else ''
            cursor.execute("""
                INSERT INTO registros (fecha, empresa, area, alcance, fuente, categoria, actividad, unidad, cantidad, factor, emision)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (fecha_limpia, empresa, (destino or 'Operaciones')[:50], 'Alcance 3', 'Residuos',
                  tipos[i], trat_raw, 'kg', cant, fac, emision))
            filas_guardadas += 1

        conn.commit()
        conn.close()
        flash(f"Se han guardado {filas_guardadas} registro(s) manuales exitosamente.", "success")
        return redirect("/residuos")

    return render_template("formulario_residuos.html", factores=get_factores_residuos())

# ================= REGISTRO MANUAL GENERAL =================
@app.route('/registro', methods=['GET', 'POST'])
def registro():
    if 'user_id' not in session:
        return redirect("/")

    if request.method == 'POST':
        empresa = session.get('empresa')
        alcance = request.form.get('alcance_oculto', 'Alcance 1')
        
        fecha_raw = request.form.get('fecha')
        if len(fecha_raw) == 7:
            fecha = f"{fecha_raw}-01"
        else:
            fecha = fecha_raw

        area = request.form.get('area')
        fuente = request.form.get('fuente')
        
        try:
            cantidad = float(str(request.form.get('cantidad', '0')).replace(',', '.'))
        except:
            cantidad = 0.0

        conn = get_db()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

        if alcance == 'Alcance 2':
            origen = request.form.get('origen_energia')
            sistema = request.form.get('sistema_elec')
            tiene_irec = request.form.get('tiene_irec', 'No')

            categoria = f"Electricidad {sistema}"
            unidad = "kWh"
            actividad = "Consumo de red eléctrica"

            fecha_dt = datetime.strptime(fecha, "%Y-%m-%d")
            anio_reg, mes_reg = fecha_dt.year, fecha_dt.month

            # Factor ubicación: buscar por año/mes exacto, si no existe usar el más reciente del mismo sistema
            cursor.execute(
                "SELECT factor_emision_avg FROM factores_electricos WHERE sistema = %s AND anio = %s AND mes = %s",
                (sistema, anio_reg, mes_reg)
            )
            res_ub = cursor.fetchone()
            if not res_ub:
                cursor.execute(
                    "SELECT factor_emision_avg FROM factores_electricos WHERE sistema = %s ORDER BY anio DESC, mes DESC LIMIT 1",
                    (sistema,)
                )
                res_ub = cursor.fetchone()
            factor_ubicacion = float(res_ub[0]) if res_ub and res_ub[0] is not None else 0.0
            emision_ubicacion = round(cantidad * factor_ubicacion, 4)

            # Factor mercado: 0 si ERNC + IREC válido, residual en caso contrario
            if origen == 'ERNC' and tiene_irec == 'Si':
                factor = 0.0
            else:
                cursor.execute(
                    "SELECT factor_emision_avg FROM factores_electricos WHERE LOWER(sistema) = 'residual' AND anio = %s AND mes = %s",
                    (anio_reg, mes_reg)
                )
                res_merc = cursor.fetchone()
                if not res_merc:
                    cursor.execute(
                        "SELECT factor_emision_avg FROM factores_electricos WHERE LOWER(sistema) = 'residual' ORDER BY anio DESC, mes DESC LIMIT 1"
                    )
                    res_merc = cursor.fetchone()
                if not res_merc:
                    # Fallback final: factor SEN más cercano
                    cursor.execute(
                        "SELECT factor_emision_avg FROM factores_electricos WHERE sistema = %s ORDER BY ABS(anio - %s) ASC, ABS(mes - %s) ASC LIMIT 1",
                        (sistema, anio_reg, mes_reg)
                    )
                    res_merc = cursor.fetchone()
                factor = float(res_merc[0]) if res_merc and res_merc[0] is not None else 0.0

            emision = round(cantidad * factor, 4)

            # Guardar certificado IREC si fue subido
            archivo_irec = request.files.get('certificado_irec')
            if tiene_irec == 'Si' and archivo_irec and archivo_irec.filename:
                cursor.execute(
                    "INSERT INTO irec_certificados (empresa, fecha_consumo, filename, contenido, fecha_subida) VALUES (%s, %s, %s, %s, %s)",
                    (empresa, fecha, archivo_irec.filename, psycopg2.Binary(archivo_irec.read()), datetime.now().strftime("%Y-%m-%d %H:%M"))
                )

        else:
            categoria = request.form.get('combustible')
            unidad = request.form.get('unidad')
            actividad = "Consumo general"
            
            if categoria == 'Otros':
                categoria = request.form.get('otro_combustible')
                try:
                    factor = float(str(request.form.get('otro_factor', '0')).replace(',', '.'))
                except:
                    factor = 0.0
            else:
                try:
                    factor = float(str(request.form.get('factor_oculto', '0')).replace(',', '.'))
                except:
                    factor = 0.0
                
            emision = cantidad * factor

        if alcance == 'Alcance 2':
            cursor.execute("""
                INSERT INTO registros (fecha, empresa, area, alcance, fuente, categoria, actividad, unidad, cantidad, factor, emision, emision_ubicacion, origen_energia, tiene_irec, sistema)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (fecha, empresa, area, alcance, fuente, categoria, actividad, unidad, cantidad, factor, emision, emision_ubicacion, origen, tiene_irec, sistema))
        else:
            cursor.execute("""
                INSERT INTO registros (fecha, empresa, area, alcance, fuente, categoria, actividad, unidad, cantidad, factor, emision)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (fecha, empresa, area, alcance, fuente, categoria, actividad, unidad, cantidad, factor, emision))
        
        conn.commit()
        conn.close()
        
        if alcance == 'Alcance 2':
            flash(f"Consumo eléctrico registrado — Sistema: {sistema} | Factor ubicación: {factor_ubicacion} | Emisión mercado: {emision} kg | Emisión ubicación: {emision_ubicacion} kg", "success")
            return redirect("/electricidad")
        flash("Registro guardado exitosamente.", "success")
        return redirect("/dashboard")
            
    conn = get_db()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cursor.execute("SELECT categoria, unidad, factor, CASE WHEN TRIM(LOWER(COALESCE(nombre_chile,''))) IN ('nan','none','') THEN NULL ELSE nombre_chile END as nombre_chile, COALESCE(año, 0) as año FROM factores ORDER BY año DESC")
    _all_factores = [dict(row) for row in cursor.fetchall()]
    conn.close()
    # Deduplicate: keep latest año per (categoria, unidad)
    _seen_cu = set()
    factores_db = []
    for _f in _all_factores:
        _key = (_f['categoria'], _f['unidad'])
        if _key not in _seen_cu:
            _seen_cu.add(_key)
            factores_db.append(_f)

    nombres_principales = ['Diésel', 'Bencina/Gasolina', 'Gas Licuado Petróleo (GLP)', 'R410A']
    datos_agrupados = {'principales': {}, 'combustibles': {}, 'refrigerantes': {}, 'otros': {}}

    for f in factores_db:
        cat = f['categoria']
        cat_lower = cat.lower()
        
        if 'electricidad' in cat_lower or 'kwh' in f['unidad'].lower() or 'sen' in cat_lower:
            continue
            
        grupo = 'otros'
        if cat in nombres_principales:
            grupo = 'principales'
        elif 'gas' in cat_lower or 'diésel' in cat_lower or 'bencina' in cat_lower or 'aceite' in cat_lower or 'petróleo' in cat_lower:
            grupo = 'combustibles'
        elif 'r4' in cat_lower or 'hfc' in cat_lower or 'cfc' in cat_lower:
            grupo = 'refrigerantes'
        
        if cat not in datos_agrupados[grupo]:
            datos_agrupados[grupo][cat] = []
        
        datos_agrupados[grupo][cat].append({'unidad': f['unidad'], 'factor': f['factor']})
            
    return render_template("registro.html", datos_factores=datos_agrupados)


# ================= IMPORTACIÓN Y EXPORTACIÓN =================
def _parse_import_file(file, factores_dict=None):
    meses_dict = {
        'enero': '01', 'febrero': '02', 'marzo': '03', 'abril': '04',
        'mayo': '05', 'junio': '06', 'julio': '07', 'agosto': '08',
        'septiembre': '09', 'octubre': '10', 'noviembre': '11', 'diciembre': '12'
    }
    try:
        df = pd.read_excel(file, sheet_name='Datos Centralizados')
    except ValueError:
        raise ValueError("El Excel debe contener una pestaña llamada 'Datos Centralizados'.")

    df.columns = df.columns.str.strip()
    df['Cantidad'] = df['Cantidad'].astype(str).str.replace(',', '.', regex=False)
    df['Cantidad'] = pd.to_numeric(df['Cantidad'], errors='coerce')

    col_factor = 'Factor emisión (kg CO₂/u)'
    if col_factor in df.columns:
        df[col_factor] = df[col_factor].astype(str).str.replace(',', '.', regex=False)
        df[col_factor] = pd.to_numeric(df[col_factor], errors='coerce').fillna(0.0)
    else:
        df[col_factor] = 0.0

    filas = []
    for index, row in df.iterrows():
        errores = []
        advertencias = []

        def get_texto(columna, default=''):
            if columna not in row: return default
            valor = row[columna]
            if pd.isna(valor) or str(valor).strip().lower() == 'nan' or str(valor).strip() == '':
                return default
            return str(valor).strip()

        # Saltar fila de descripción (texto largo en Mes) y fila de ejemplo del template
        mes_raw = get_texto('Mes', '')
        if len(mes_raw) > 20:
            continue
        if (get_texto('Identificador', '') == 'GEN-01' and
                get_texto('Tipo de Combustible', '') == 'Diésel' and
                get_texto('Sucursal', '') == 'Sede Central'):
            continue

        mes_texto = mes_raw.lower()
        anio_str = get_texto('Año', str(datetime.now().year)).replace('.0', '')

        # Validar año
        try:
            anio_int = int(float(anio_str))
            if anio_int < 2000 or anio_int > 2035:
                errores.append(f"Año fuera de rango: {anio_str}")
        except (ValueError, TypeError):
            errores.append(f"Año inválido: '{anio_str}'")
            anio_int = datetime.now().year
        anio = str(anio_int)

        if not mes_texto:
            errores.append("Mes vacío")
        elif mes_texto not in meses_dict:
            errores.append(f"Mes inválido: '{mes_texto}'")
        mes_num = meses_dict.get(mes_texto, '01')
        fecha_sql = f"{anio}-{mes_num}-01"

        excel_level1 = get_texto('Level 1', '')
        excel_level2 = get_texto('Level 2', '')

        fuente = 'Desconocida'
        if excel_level1.lower() == 'combustibles':
            if 'fija' in excel_level2.lower(): fuente = 'Combustión Fija'
            elif 'móvil' in excel_level2.lower() or 'movil' in excel_level2.lower(): fuente = 'Combustible Móvil'
            else: fuente = 'Combustión Fija'
        elif excel_level1.lower() == 'electricidad': fuente = 'Electricidad'
        elif excel_level1.lower() == 'refrigerantes': fuente = 'Refrigerantes'
        elif excel_level1.lower() == 'residuos': fuente = 'Residuos'
        elif excel_level1 == '':
            errores.append("Level 1 vacío")
        else:
            errores.append(f"Level 1 desconocido: '{excel_level1}'")

        if fuente in ('Combustión Fija', 'Combustible Móvil', 'Combustión Estacionaria', 'Refrigerantes', 'Fugas de Refrigerantes'):
            alcance = 'Alcance 1'
        elif fuente == 'Electricidad':
            alcance = 'Alcance 2'
        elif fuente == 'Residuos':
            alcance = 'Alcance 3'
        else:
            alcance = get_texto('Scope', 'No definido')
        area = get_texto('Sucursal', 'General')
        categoria = get_texto('Tipo de Combustible', '')
        actividad = get_texto('Tipo Unidad de Consumo', '')
        identificador = get_texto('Identificador', '')
        unidad = get_texto('Unidad de Medida', 'N/A')

        if fuente == 'Electricidad':
            if categoria == '': categoria = 'Red Eléctrica'
            if actividad == '': actividad = identificador if identificador != '' else 'Consumo General'

        if categoria == '': categoria = 'Desconocida'
        if actividad == '': actividad = 'No especificada'

        # Validar cantidad (NaN o vacío se trata como 0 para permitir registros mensuales sin consumo)
        if pd.isna(row['Cantidad']):
            cantidad = 0.0
        else:
            cantidad = float(row['Cantidad'])
            if cantidad < 0:
                errores.append(f"Cantidad negativa ({cantidad})")

        # Factor: columna Excel → auto-lookup en catálogo → 0
        factor_excel = float(row[col_factor]) if not pd.isna(row[col_factor]) else 0.0
        if factor_excel < 0:
            errores.append(f"Factor negativo ({factor_excel})")
            factor_excel = 0.0

        factor = factor_excel
        factor_origen = 'excel'
        if factor == 0.0 and factores_dict and categoria:
            lookup = factores_dict.get((categoria.lower(), unidad.lower()))
            if lookup:
                factor = lookup
                factor_origen = 'auto'
                advertencias.append(f"Factor completado automáticamente desde catálogo ({factor} kg CO₂/{unidad})")

        emision = cantidad * factor

        filas.append({
            'row_num': index + 2,
            'fecha': fecha_sql,
            'mes': mes_raw,
            'anio': anio,
            'fuente': fuente,
            'alcance': alcance,
            'area': area,
            'categoria': categoria,
            'actividad': actividad,
            'unidad': unidad,
            'cantidad': cantidad,
            'factor': factor,
            'factor_origen': factor_origen,
            'emision': emision,
            'valid': len(errores) == 0,
            'error': '; '.join(errores) if errores else None,
            'advertencia': '; '.join(advertencias) if advertencias else None
        })

    return filas


@app.route("/descargar_plantilla")
def descargar_plantilla():
    if 'user_id' not in session:
        return redirect("/")

    columnas = [
        'Mes', 'Año', 'Level 1', 'Level 2', 'Scope', 'Sucursal',
        'Tipo de Combustible', 'Tipo Unidad de Consumo', 'Identificador',
        'Unidad de Medida', 'Cantidad', 'Factor emisión (kg CO₂/u)'
    ]
    descripciones = [
        'Nombre del mes (Enero, Febrero…)', 'Año (ej: 2025)',
        'Combustibles / Electricidad / Refrigerantes / Residuos',
        'Combustión Fija / Combustión Móvil (solo Combustibles)',
        'Alcance 1 / Alcance 2 / Alcance 3',
        'Nombre de la sucursal o sede',
        'Tipo de combustible o gas refrigerante',
        'Tipo de consumo o proceso',
        'Código o identificador del equipo/medidor',
        'Litros, kWh, kg, etc.',
        'Cantidad consumida (número)',
        'Factor de emisión oficial (kg CO₂ por unidad)'
    ]
    ejemplo = ['Enero', 2025, 'Combustibles', 'Combustión Fija', 'Alcance 1', 'Sede Central', 'Diésel', 'Generador', 'GEN-01', 'Litros', 1500.5, 2.68]

    conn = get_db()
    df_factores = pd.read_sql_query(
        "SELECT categoria as \"Categoría\", unidad as \"Unidad\", factor as \"Factor Oficial\" FROM factores ORDER BY categoria",
        conn
    )
    conn.close()

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        workbook = writer.book

        header_fmt = workbook.add_format({
            'bold': True, 'bg_color': '#1E40AF', 'font_color': '#FFFFFF',
            'border': 1, 'text_wrap': True, 'valign': 'vcenter', 'align': 'center'
        })
        desc_fmt = workbook.add_format({
            'italic': True, 'bg_color': '#DBEAFE', 'font_color': '#1E40AF',
            'border': 1, 'text_wrap': True, 'font_size': 9
        })
        example_fmt = workbook.add_format({
            'bg_color': '#F0FDF4', 'font_color': '#166534', 'border': 1, 'bold': True
        })
        num_fmt = workbook.add_format({
            'bg_color': '#F0FDF4', 'font_color': '#166534', 'border': 1,
            'bold': True, 'num_format': '#,##0.00'
        })

        worksheet = workbook.add_worksheet('Datos Centralizados')
        writer.sheets['Datos Centralizados'] = worksheet

        col_widths = [12, 8, 18, 22, 12, 18, 24, 24, 16, 16, 12, 26]
        for col, (name, width) in enumerate(zip(columnas, col_widths)):
            worksheet.write(0, col, name, header_fmt)
            worksheet.set_column(col, col, width)
        for col, desc in enumerate(descripciones):
            worksheet.write(1, col, desc, desc_fmt)
        for col, val in enumerate(ejemplo):
            fmt = num_fmt if col in (1, 10, 11) else example_fmt
            worksheet.write(2, col, val, fmt)

        worksheet.set_row(0, 30)
        worksheet.set_row(1, 42)
        worksheet.set_row(2, 20)
        worksheet.freeze_panes(1, 0)

        if not df_factores.empty:
            df_factores.to_excel(writer, sheet_name='Catálogo Oficial', index=False)
            ws2 = writer.sheets['Catálogo Oficial']
            for ci, col_name in enumerate(df_factores.columns):
                ws2.write(0, ci, col_name, header_fmt)
            ws2.set_column('A:A', 30)
            ws2.set_column('B:B', 15)
            ws2.set_column('C:C', 18)

    output.seek(0)
    return send_file(output, download_name="Plantilla_Masiva_GreenTrack.xlsx", as_attachment=True)


@app.route("/importar", methods=["GET", "POST"])
def importar_registros():
    if 'user_id' not in session: return redirect("/")

    if request.method == "POST":
        if 'archivo' not in request.files or request.files['archivo'].filename == '':
            flash("No se seleccionó ningún archivo", "error")
            return redirect(request.url)

        file = request.files['archivo']
        if not file or not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
            flash("Formato no válido. Debe ser .xlsx", "error")
            return redirect(request.url)

        try:
            conn = get_db()
            cur = conn.cursor()
            cur.execute("SELECT categoria, unidad, factor FROM factores")
            factores_dict = {(r[0].lower(), r[1].lower()): float(r[2]) for r in cur.fetchall()}
            conn.close()
        except Exception:
            factores_dict = {}

        try:
            filas = _parse_import_file(file, factores_dict)
        except ValueError as e:
            flash(str(e), "error")
            return redirect(request.url)
        except Exception as e:
            flash(f"Error técnico al procesar: {str(e)}", "error")
            return redirect(request.url)

        total = len(filas)
        validos = sum(1 for f in filas if f['valid'])
        errores = total - validos
        total_emision = sum(f['emision'] for f in filas if f['valid'])

        from collections import defaultdict
        _fc = defaultdict(lambda: {'count': 0, 'emision': 0.0})
        for f in filas:
            if f['valid']:
                _fc[f['fuente']]['count'] += 1
                _fc[f['fuente']]['emision'] += f['emision']
        resumen_fuentes = [
            {'fuente': k, 'count': v['count'], 'emision': v['emision']}
            for k, v in sorted(_fc.items(), key=lambda x: -x[1]['emision'])
        ]

        return render_template("importar_preview.html",
            filas=filas,
            filas_json=json.dumps(filas),
            total=total,
            validos=validos,
            errores=errores,
            total_emision=total_emision,
            resumen_fuentes=resumen_fuentes
        )

    return render_template("importar.html")


@app.route("/importar/confirmar", methods=["POST"])
def importar_confirmar():
    if 'user_id' not in session: return redirect("/")
    empresa = session.get('empresa')

    filas_json = request.form.get('filas_json', '[]')
    try:
        filas = json.loads(filas_json)
    except Exception:
        flash("Error al procesar los datos. Intente nuevamente.", "error")
        return redirect("/importar")

    filas_validas = [f for f in filas if f.get('valid')]
    if not filas_validas:
        flash("No hay registros válidos para importar.", "error")
        return redirect("/importar")

    try:
        conn = get_db()
        cursor = conn.cursor()
        for fila in filas_validas:
            cursor.execute("""
                INSERT INTO registros (fecha, empresa, area, alcance, fuente, categoria, actividad, unidad, cantidad, factor, emision)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (fila['fecha'], empresa, fila['area'], fila['alcance'], fila['fuente'],
                  fila['categoria'], fila['actividad'], fila['unidad'],
                  fila['cantidad'], fila['factor'], fila['emision']))
        conn.commit()
        conn.close()

        from collections import defaultdict
        _fc = defaultdict(lambda: {'count': 0, 'emision': 0.0})
        for f in filas_validas:
            _fc[f['fuente']]['count'] += 1
            _fc[f['fuente']]['emision'] += f['emision']

        session['import_result'] = {
            'guardados': len(filas_validas),
            'total_emision': sum(f['emision'] for f in filas_validas),
            'por_fuente': [
                {'fuente': k, 'count': v['count'], 'emision': v['emision']}
                for k, v in sorted(_fc.items(), key=lambda x: -x[1]['emision'])
            ]
        }
        return redirect("/importar/resultado")

    except Exception as e:
        flash(f"Error al guardar los datos: {str(e)}", "error")
        return redirect("/importar")


@app.route("/importar/resultado")
def importar_resultado():
    if 'user_id' not in session: return redirect("/")
    result = session.pop('import_result', None)
    if not result:
        return redirect("/importar")
    return render_template("importar_resultado.html",
        guardados=result['guardados'],
        total_emision=result['total_emision'],
        por_fuente=result['por_fuente']
    )

@app.route("/exportar", methods=["GET"])
def exportar():
    if 'user_id' not in session: return redirect("/")
    empresa = session.get('empresa')
    conn = get_db()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cursor.execute("SELECT DISTINCT SUBSTRING(fecha::text,1,4) as anio FROM registros WHERE empresa=%s AND fecha IS NOT NULL ORDER BY anio DESC", (empresa,))
    anios = [r[0] for r in cursor.fetchall()]
    cursor.execute("SELECT DISTINCT alcance FROM registros WHERE empresa=%s AND alcance IS NOT NULL ORDER BY alcance", (empresa,))
    alcances = [r[0] for r in cursor.fetchall()]
    cursor.execute("SELECT DISTINCT fuente FROM registros WHERE empresa=%s AND fuente IS NOT NULL ORDER BY fuente", (empresa,))
    fuentes = [r[0] for r in cursor.fetchall()]
    conn.close()
    return render_template("exportar.html", empresa=empresa, anios=anios, alcances=alcances, fuentes=fuentes)

@app.route("/exportar_completo")
def exportar_completo():
    return exportar_avanzado()

@app.route("/exportar_avanzado")
def exportar_avanzado():
    if 'user_id' not in session:
        return redirect("/")

    empresa = session.get('empresa')
    anio        = request.args.get('anio', '').strip()
    alcances_sel = request.args.getlist('alcance')
    fuentes_sel  = request.args.getlist('fuente')
    fecha_inicio = request.args.get('fecha_inicio', '').strip()
    fecha_fin    = request.args.get('fecha_fin', '').strip()

    conn = get_db()
    where = ["empresa = %s"]
    params = [empresa]

    if anio:
        where.append("SUBSTRING(fecha::text,1,4) = %s"); params.append(anio)
    if alcances_sel:
        where.append(f"alcance IN %s"); params.append(tuple(alcances_sel))
    if fuentes_sel:
        where.append(f"fuente IN %s"); params.append(tuple(fuentes_sel))
    if fecha_inicio:
        where.append("fecha >= %s"); params.append(fecha_inicio)
    if fecha_fin:
        where.append("fecha <= %s"); params.append(fecha_fin)

    where_sql = " AND ".join(where)
    df = pd.read_sql_query(
        f"SELECT fecha, area, alcance, fuente, categoria, actividad, unidad, cantidad, factor, emision "
        f"FROM registros WHERE {where_sql} ORDER BY fecha DESC, id DESC",
        conn, params=tuple(params))
    conn.close()

    if df.empty:
        flash("No hay registros que coincidan con los filtros seleccionados.", "warning")
        return redirect("/exportar")

    df.rename(columns={
        'fecha': 'Fecha', 'area': 'Área / Instalación', 'alcance': 'Alcance',
        'fuente': 'Fuente', 'categoria': 'Categoría', 'actividad': 'Actividad',
        'unidad': 'Unidad', 'cantidad': 'Cantidad', 'factor': 'Factor (kg CO₂/u)',
        'emision': 'Emisión (kg CO₂e)'
    }, inplace=True)

    from openpyxl.styles import Font, PatternFill, Alignment
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Registros', index=False)
        ws = writer.sheets['Registros']
        header_fill = PatternFill("solid", fgColor="1E40AF")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 20

    output.seek(0)
    sufijo = f"_{anio}" if anio else "_completo"
    nombre = f"Datos_GreenTrack_{empresa.replace(' ','_')}{sufijo}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(output, download_name=nombre, as_attachment=True)


# ================= RUTAS ADMIN =================
def get_admin_stats():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) FROM usuarios WHERE es_admin = 0")
    total_emp = cursor.fetchone()[0] or 0
    cursor.execute("SELECT COUNT(*) FROM registros")
    total_reg = cursor.fetchone()[0] or 0
    cursor.execute("SELECT SUM(emision) FROM registros")
    total_em = cursor.fetchone()[0] or 0
    try:
        cursor.execute("SELECT COUNT(*) FROM pending_pdf_uploads WHERE estado = 'pendiente'")
        pendientes = cursor.fetchone()[0] or 0
    except Exception:
        pendientes = 0
    conn.close()
    return total_emp, total_reg, total_em, pendientes

@app.route("/admin/dashboard")
def admin_dashboard():
    if 'user_id' not in session or session.get('es_admin') != 1: return redirect("/")
    t_emp, t_reg, t_em, t_pend = get_admin_stats()

    conn = get_db()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cursor.execute("SELECT id, empresa, email, contacto, fecha_registro FROM usuarios WHERE es_admin = 0 ORDER BY fecha_registro DESC LIMIT 5")
    ultimas = [dict(row) for row in cursor.fetchall()]
    cursor.execute("SELECT id, empresa FROM usuarios WHERE es_admin = 0 ORDER BY empresa")
    empresas_filtro = [dict(row) for row in cursor.fetchall()]

    cursor.execute("""
        SELECT empresa, SUBSTRING(fecha::text,1,4) AS anio, SUM(emision) AS total
        FROM registros WHERE empresa IS NOT NULL AND fecha IS NOT NULL AND LENGTH(fecha::text) >= 4
        GROUP BY empresa, anio
    """)
    em_map = {}
    for row in cursor.fetchall():
        emp, anio_val, total = row['empresa'], row['anio'], float(row['total'] or 0)
        if emp and anio_val:
            if emp not in em_map: em_map[emp] = {}
            em_map[emp][anio_val] = em_map[emp].get(anio_val, 0) + total

    conn.close()

    all_empresas = sorted(em_map.keys())
    all_anios = sorted({a for d in em_map.values() for a in d.keys()}, reverse=True)
    chart_data = {"Todos": [round(sum(em_map.get(e, {}).values()), 2) for e in all_empresas]}
    for anio in all_anios:
        chart_data[anio] = [round(em_map.get(e, {}).get(anio, 0), 2) for e in all_empresas]

    top_emisores = sorted(
        [{'empresa': e, 'total': round(sum(v.values()), 2)} for e, v in em_map.items()],
        key=lambda x: -x['total']
    )[:6]

    return render_template("admin.html",
        total_empresas=t_emp, total_registros=t_reg, total_emisiones=t_em, total_pendientes=t_pend,
        ultimas_empresas=ultimas, empresas=empresas_filtro, admin_section="dashboard",
        chart_empresas=all_empresas, chart_data=chart_data, chart_anios=all_anios,
        top_emisores=top_emisores)

@app.route("/admin/empresas")
def admin_empresas():
    if session.get('es_admin') != 1: return redirect("/")
    t_emp, t_reg, t_em, t_pend = get_admin_stats()
    conn = get_db()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cursor.execute("SELECT id, empresa, email, contacto, rut, fecha_registro FROM usuarios WHERE es_admin = 0 ORDER BY empresa")
    empresas = [dict(row) for row in cursor.fetchall()]
    cursor.execute("SELECT empresa, COUNT(*) as cnt, SUM(emision) as total FROM registros GROUP BY empresa")
    em_stats = {}
    for r in cursor.fetchall():
        if r['empresa']:
            em_stats[r['empresa']] = {'cnt': int(r['cnt'] or 0), 'total': float(r['total'] or 0)}
    cursor.execute("SELECT DISTINCT SUBSTRING(fecha::text,1,4) as anio FROM registros WHERE fecha IS NOT NULL ORDER BY anio DESC")
    anios_export = [r[0] for r in cursor.fetchall()]
    cursor.execute("SELECT DISTINCT fuente FROM registros WHERE fuente IS NOT NULL ORDER BY fuente")
    fuentes_export = [r[0] for r in cursor.fetchall()]
    conn.close()

    return render_template("admin.html", empresas=empresas, em_stats=em_stats, admin_section="empresas",
        total_empresas=t_emp, total_registros=t_reg, total_emisiones=t_em, total_pendientes=t_pend,
        anios_export=anios_export, fuentes_export=fuentes_export)

@app.route("/admin/factores", methods=["GET", "POST"])
def admin_factores():
    if session.get('es_admin') != 1: return redirect("/")
    t_emp, t_reg, t_em, t_pend = get_admin_stats()
    conn = get_db()
    if request.method == "POST":
        try:
            año_val = int(request.form.get("año") or 0)
            conn.cursor().execute("""
                INSERT INTO factores (categoria, unidad, factor, año)
                VALUES (%s, %s, %s, %s)
                ON CONFLICT (categoria, unidad, año) DO UPDATE SET factor = EXCLUDED.factor
            """, (request.form.get("categoria"), request.form.get("unidad"), float(request.form.get("factor")), año_val))
            conn.commit()
            flash("Factor guardado exitosamente", "success")
        except Exception as e:
            flash(f"Error al guardar: {e}", "error")
            conn.rollback()

    df = pd.read_sql_query("SELECT * FROM factores ORDER BY año DESC, categoria, unidad", conn)
    # pandas convierte NULL a float('nan') — convertir a None para que Jinja2 los trate como falsy
    factores_list = [
        {k: (None if (isinstance(v, float) and pd.isna(v)) else v) for k, v in row.items()}
        for row in df.to_dict('records')
    ]
    cur_elec = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur_elec.execute("SELECT anio, mes, sistema, factor_emision_avg FROM factores_electricos ORDER BY sistema, anio, mes")
    factores_elec = [dict(r) for r in cur_elec.fetchall()]
    resumen_elec = {}
    for f in factores_elec:
        key = (f['sistema'], f['anio'])
        resumen_elec[key] = resumen_elec.get(key, 0) + 1
    conn.close()
    return render_template("admin.html", factores=factores_list, factores_elec=factores_elec, resumen_elec=resumen_elec, admin_section="factores",
        total_empresas=t_emp, total_registros=t_reg, total_emisiones=t_em, total_pendientes=t_pend)

@app.route("/admin/editar_empresa/<int:empresa_id>", methods=["POST"])
def admin_editar_empresa(empresa_id):
    if session.get('es_admin') != 1: return redirect("/")
    nombre = request.form.get("empresa", "").strip()
    email = request.form.get("email", "").strip()
    contacto = request.form.get("contacto", "").strip()
    rut = request.form.get("rut", "").strip()
    if not nombre or not email:
        flash("Nombre y email son obligatorios.", "error")
        return redirect(url_for('admin_empresas'))
    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT empresa FROM usuarios WHERE id = %s AND es_admin = 0", (empresa_id,))
        row = cursor.fetchone()
        if not row:
            flash("Empresa no encontrada.", "error")
            conn.close()
            return redirect(url_for('admin_empresas'))
        nombre_anterior = row[0]
        cursor.execute("""
            UPDATE usuarios SET empresa = %s, email = %s, contacto = %s, rut = %s
            WHERE id = %s AND es_admin = 0
        """, (nombre, email, contacto, rut, empresa_id))
        if nombre != nombre_anterior:
            for tabla in ['registros', 'combustible_movil', 'vehiculos', 'pdf_uploads',
                          'agua_consumo', 'agua_cuencas', 'agua_afluentes', 'agua_costos',
                          'irec_certificados', 'configuracion', 'energeticos_empresa']:
                cursor.execute(f"UPDATE {tabla} SET empresa = %s WHERE empresa = %s", (nombre, nombre_anterior))
        conn.commit()
        flash(f"Empresa '{nombre}' actualizada correctamente.", "success")
    except psycopg2.IntegrityError:
        conn.rollback()
        flash("Error: ese email ya está en uso por otra cuenta.", "error")
    except Exception as e:
        conn.rollback()
        flash(f"Error al editar: {e}", "danger")
    finally:
        conn.close()
    return redirect(url_for('admin_empresas'))


@app.route("/admin/resetear_password/<int:empresa_id>", methods=["POST"])
def admin_resetear_password(empresa_id):
    if session.get('es_admin') != 1: return redirect("/")
    nueva = request.form.get("nueva_password", "").strip()
    if len(nueva) < 6:
        flash("La contraseña debe tener al menos 6 caracteres.", "error")
        return redirect(url_for('admin_empresas'))
    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT empresa FROM usuarios WHERE id = %s AND es_admin = 0", (empresa_id,))
        row = cursor.fetchone()
        if not row:
            flash("Empresa no encontrada.", "error")
            conn.close()
            return redirect(url_for('admin_empresas'))
        cursor.execute("UPDATE usuarios SET password = %s WHERE id = %s", (hash_password(nueva), empresa_id))
        conn.commit()
        flash(f"Contraseña de '{row[0]}' reseteada correctamente.", "success")
    except Exception as e:
        conn.rollback()
        flash(f"Error: {e}", "danger")
    finally:
        conn.close()
    return redirect(url_for('admin_empresas'))


@app.route("/admin/eliminar_empresa/<int:empresa_id>", methods=["POST"])
def admin_eliminar_empresa(empresa_id):
    if session.get('es_admin') != 1: return redirect("/")
    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT empresa FROM usuarios WHERE id = %s AND es_admin = 0", (empresa_id,))
        row = cursor.fetchone()
        if not row:
            flash("Empresa no encontrada.", "error")
            conn.close()
            return redirect(url_for('admin_empresas'))
        nombre = row[0]
        for tabla in ['registros', 'combustible_movil', 'vehiculos', 'pdf_uploads',
                      'agua_consumo', 'agua_cuencas', 'agua_afluentes', 'agua_costos',
                      'irec_certificados', 'configuracion', 'energeticos_empresa']:
            cursor.execute(f"DELETE FROM {tabla} WHERE empresa = %s", (nombre,))
        cursor.execute("DELETE FROM usuarios WHERE id = %s AND es_admin = 0", (empresa_id,))
        conn.commit()
        flash(f"Empresa '{nombre}' y todos sus datos han sido eliminados.", "success")
    except Exception as e:
        conn.rollback()
        flash(f"Error al eliminar: {e}", "danger")
    finally:
        conn.close()
    return redirect(url_for('admin_empresas'))


@app.route("/admin/crear_empresa", methods=["POST"])
def admin_crear_empresa():
    if session.get('es_admin') != 1: return redirect("/")
    empresa, email = request.form.get("empresa"), request.form.get("email")
    password = hash_password(request.form.get("password"))
    contacto, rut = request.form.get("contacto"), request.form.get("rut")

    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            INSERT INTO usuarios (empresa, email, password, contacto, rut, fecha_registro, es_admin)
            VALUES (%s, %s, %s, %s, %s, %s, 0)
        """, (empresa, email, password, contacto, rut, datetime.now().strftime("%Y-%m-%d %H:%M")))
        conn.commit()
        flash(f"Empresa '{empresa}' creada exitosamente.", "success")
    except psycopg2.IntegrityError:
        conn.rollback()
        flash("Error: El correo electrónico ya está en uso.", "error")
    finally: conn.close()
    return redirect(url_for('admin_empresas'))

@app.route("/admin/debug_excel", methods=["POST"])
def admin_debug_excel():
    if session.get('es_admin') != 1: return redirect("/")
    file = request.files.get('archivo_factores')
    if not file or file.filename == '':
        return "No se subió archivo", 400
    df = pd.read_excel(file, sheet_name='Equivalencias', header=None, engine='openpyxl')
    html = ['<style>table{border-collapse:collapse}td,th{border:1px solid #ccc;padding:4px 8px;font-size:12px;white-space:nowrap}</style>']
    html.append(f'<p><b>Filas totales:</b> {len(df)} &nbsp; <b>Columnas:</b> {len(df.columns)}</p>')
    html.append('<table><thead><tr><th>Fila</th>')
    for c in df.columns:
        html.append(f'<th>Col {c}</th>')
    html.append('</tr></thead><tbody>')
    for i in range(min(12, len(df))):
        html.append(f'<tr><td><b>{i}</b></td>')
        for val in df.iloc[i]:
            cell = str(val) if pd.notna(val) else '<span style="color:#ccc">—</span>'
            html.append(f'<td>{cell}</td>')
        html.append('</tr>')
    html.append('</tbody></table>')
    return ''.join(html)

@app.route("/admin/eliminar_factor/<categoria>/<unidad>/<int:anio>", defaults={'tratamiento': ''})
@app.route("/admin/eliminar_factor/<categoria>/<unidad>/<int:anio>/<tratamiento>")
def admin_eliminar_factor(categoria, unidad, anio, tratamiento):
    if session.get('es_admin') != 1: return redirect("/")
    conn = get_db()
    cursor = conn.cursor()
    if tratamiento:
        cursor.execute(
            "DELETE FROM factores WHERE categoria = %s AND unidad = %s AND año = %s AND tratamiento = %s",
            (categoria, unidad, anio, tratamiento))
    else:
        cursor.execute(
            "DELETE FROM factores WHERE categoria = %s AND unidad = %s AND año = %s AND (tratamiento IS NULL OR tratamiento = '')",
            (categoria, unidad, anio))
    conn.commit()
    conn.close()
    flash("Factor eliminado.", "info")
    return redirect(url_for('admin_factores'))

@app.route("/admin/cargar_factores", methods=["POST"])
def cargar_factores():
    if session.get('es_admin') != 1: return redirect("/")
    if 'archivo_factores' not in request.files or request.files['archivo_factores'].filename == '':
        flash("No se subió ningún archivo", "danger")
        return redirect(request.referrer)
        
    file = request.files['archivo_factores']
    try:
        año_factores = int(request.form.get("año_factores") or 0)
        df = pd.read_excel(file, sheet_name='Equivalencias', header=None, engine='openpyxl')
        conn = get_db()
        cursor = conn.cursor()
        if año_factores:
            cursor.execute("DELETE FROM factores WHERE año = %s", (año_factores,))
        else:
            cursor.execute("DELETE FROM factores WHERE año = 0")
        actualizados = 0

        def safe_get(row_data, idx):
            if idx >= len(row_data):
                return None
            val = row_data.iloc[idx]
            s = str(val).strip()
            return s if pd.notna(val) and s.lower() not in ('nan', 'none', '') else None

        def safe_float(val):
            if val is None: return None
            try: return float(str(val).replace(',', '.'))
            except: return None

        # Índices fijos confirmados para la hoja Equivalencias
        # Col 2=cat combustible, 5=unidad, 7=factor comb
        # Col 12=cat refrigerante, 14=factor ref
        # Col 19=categoría DEFRA residuo, 21=tratamiento, 22=equivalencia Chile, 23=factor residuo
        IDX_DEFRA   = 19
        IDX_TRAT    = 21
        IDX_EQUIV   = 22   # "Equivalencia en Chile"
        IDX_FE_RES  = 23

        residuos_map = {}  # {defra_name: {'fe', 'nombre_chile', 'trat'}}

        for index, row in df.iterrows():
            if index < 3:
                continue

            # Combustibles
            cat_comb = safe_get(row, 2)
            uni_comb = safe_get(row, 5) or 'N/A'
            fe_comb  = safe_float(safe_get(row, 7))
            if cat_comb and fe_comb is not None:
                cursor.execute("""
                    INSERT INTO factores (categoria, unidad, factor, año) VALUES (%s, %s, %s, %s)
                    ON CONFLICT (categoria, unidad, año, COALESCE(tratamiento, ''))
                    DO UPDATE SET factor = EXCLUDED.factor
                """, (cat_comb, uni_comb, fe_comb, año_factores))
                actualizados += 1

            # Refrigerantes
            cat_ref = safe_get(row, 12)
            fe_ref  = safe_float(safe_get(row, 14))
            if cat_ref and fe_ref is not None:
                cursor.execute("""
                    INSERT INTO factores (categoria, unidad, factor, año) VALUES (%s, %s, %s, %s)
                    ON CONFLICT (categoria, unidad, año, COALESCE(tratamiento, ''))
                    DO UPDATE SET factor = EXCLUDED.factor
                """, (cat_ref, 'kg', fe_ref, año_factores))
                actualizados += 1

            # Residuos: guardar un registro por (DEFRA, tratamiento) — no colapsar al máximo
            cat_defra   = safe_get(row, IDX_DEFRA)
            equiv_chile = safe_get(row, IDX_EQUIV)
            trat_esp    = safe_get(row, IDX_TRAT) or ''
            fe_res      = safe_float(safe_get(row, IDX_FE_RES))

            if cat_defra and fe_res is not None and fe_res > 0:
                key = (cat_defra, trat_esp)
                existing = residuos_map.get(key)
                if existing is None:
                    residuos_map[key] = {'fe': fe_res, 'nombre_chile': equiv_chile, 'trat': trat_esp}
                else:
                    if equiv_chile and not existing['nombre_chile']:
                        existing['nombre_chile'] = equiv_chile

        for (defra, trat_key), data in residuos_map.items():
            cursor.execute("""
                INSERT INTO factores (categoria, unidad, factor, nombre_chile, tratamiento, año)
                VALUES (%s, %s, %s, %s, %s, %s)
                ON CONFLICT (categoria, unidad, año, COALESCE(tratamiento, ''))
                DO UPDATE
                    SET factor = EXCLUDED.factor,
                        nombre_chile = EXCLUDED.nombre_chile
            """, (defra, 'tonne', data['fe'], data['nombre_chile'], trat_key or None, año_factores))
            actualizados += 1

        conn.commit()
        conn.close()
        flash(f"¡Sincronización exitosa! Se actualizaron {actualizados} factores.", "success")
    except Exception as e:
        flash(f"Error técnico al leer el Excel. Detalle: {str(e)}", "danger")
    return redirect(request.referrer)

@app.route("/admin/preview_electricidad", methods=["POST"])
def preview_electricidad():
    if session.get('es_admin') != 1: return redirect("/")
    archivo = request.files.get("archivo_electricidad")
    if not archivo or archivo.filename == "":
        flash("No se seleccionó ningún archivo.", "danger")
        return redirect(url_for('admin_factores'))
    try:
        df = pd.read_excel(archivo, sheet_name='Factores eléctricos')
        filas = []
        errores = []
        for idx, row in df.iterrows():
            if pd.notna(row.iloc[0]) and pd.notna(row.iloc[3]):
                try:
                    filas.append({
                        'anio': int(row.iloc[0]),
                        'mes': int(row.iloc[1]),
                        'sistema': str(row.iloc[2]).strip(),
                        'factor': float(row.iloc[3])
                    })
                except (ValueError, TypeError):
                    errores.append(f"Fila {idx + 2}: valor inválido — {list(row[:4])}")
        sistemas = sorted(set(f['sistema'] for f in filas))
        años = sorted(set(f['anio'] for f in filas))
        return render_template("admin_preview_electricidad.html",
                               filas=filas, errores=errores,
                               sistemas=sistemas, años=años,
                               datos_json=json.dumps(filas))
    except Exception as e:
        flash(f"Error al leer el Excel: {str(e)}", "danger")
        return redirect(url_for('admin_factores'))

@app.route("/admin/cargar_electricidad", methods=["POST"])
def cargar_electricidad():
    if session.get('es_admin') != 1: return redirect("/")
    datos_json = request.form.get('datos_json')
    if not datos_json:
        # Carga directa legacy (sin preview)
        archivo = request.files.get("archivo_electricidad")
        if not archivo or archivo.filename == "":
            flash("No se seleccionó ningún archivo.", "danger")
            return redirect(request.referrer)
        try:
            df = pd.read_excel(archivo, sheet_name='Factores eléctricos')
            filas = []
            for _, row in df.iterrows():
                if pd.notna(row.iloc[0]) and pd.notna(row.iloc[3]):
                    try:
                        filas.append({'anio': int(row.iloc[0]), 'mes': int(row.iloc[1]),
                                      'sistema': str(row.iloc[2]).strip(), 'factor': float(row.iloc[3])})
                    except (ValueError, TypeError): continue
            datos_json = json.dumps(filas)
        except Exception as e:
            flash(f"Error al leer el Excel: {str(e)}", "danger")
            return redirect(request.referrer)
    try:
        filas = json.loads(datos_json)
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM factores_electricos")
        for f in filas:
            cursor.execute(
                "INSERT INTO factores_electricos (anio, mes, sistema, factor_emision_avg) VALUES (%s, %s, %s, %s)",
                (f['anio'], f['mes'], f['sistema'], f['factor'])
            )
        conn.commit()
        conn.close()
        flash(f"¡{len(filas)} factores eléctricos cargados correctamente!", "success")
    except Exception as e:
        flash(f"Error al guardar: {str(e)}", "danger")
    return redirect(url_for('admin_factores'))

@app.route("/admin/empresa/<string:nombre_empresa>")
def admin_detalle_empresa(nombre_empresa):
    if session.get('es_admin') != 1: return redirect("/")
    conn = get_db()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    
    cursor.execute("SELECT COUNT(*) as total_reg, SUM(emision) as total_emi FROM registros WHERE empresa = %s", (nombre_empresa,))
    kpis = cursor.fetchone()

    cursor.execute("SELECT fuente, SUM(emision) as total FROM registros WHERE empresa = %s GROUP BY fuente", (nombre_empresa,))
    datos_fuente = [dict(row) for row in cursor.fetchall()]

    cursor.execute("SELECT SUBSTRING(fecha, 1, 7) as mes, SUM(emision) as total FROM registros WHERE empresa = %s GROUP BY mes ORDER BY mes", (nombre_empresa,))
    datos_mes = [dict(row) for row in cursor.fetchall()]

    cursor.execute("SELECT DISTINCT SUBSTRING(fecha::text,1,4) as anio FROM registros WHERE empresa = %s AND fecha IS NOT NULL ORDER BY anio DESC", (nombre_empresa,))
    anios_disponibles = [r[0] for r in cursor.fetchall()]

    cursor.execute("SELECT DISTINCT fuente FROM registros WHERE empresa = %s AND fuente IS NOT NULL ORDER BY fuente", (nombre_empresa,))
    fuentes_disponibles = [r[0] for r in cursor.fetchall()]
    conn.close()

    return render_template("admin_detalle.html", empresa=nombre_empresa, kpis=kpis,
                           datos_fuente=datos_fuente, datos_mes=datos_mes,
                           anios_disponibles=anios_disponibles, fuentes_disponibles=fuentes_disponibles)

@app.route("/admin/exportar/<string:nombre_empresa>")
def exportar_datos_empresa(nombre_empresa):
    if session.get('es_admin') != 1: return redirect("/")
    anio    = request.args.get('anio', 'Todos')
    fuente  = request.args.get('fuente', '').strip()
    alcance = request.args.get('alcance', '').strip()

    where = ["empresa = %s"]
    params = [nombre_empresa]
    if anio != 'Todos':
        where.append("SUBSTRING(fecha::text,1,4) = %s"); params.append(anio)
    if fuente:
        where.append("fuente = %s"); params.append(fuente)
    if alcance:
        _case = ("CASE WHEN fuente IN ('Combustión Fija','Combustible Móvil','Combustión Estacionaria','Refrigerantes','Fugas de Refrigerantes') THEN 'Alcance 1'"
                 " WHEN fuente='Electricidad' THEN 'Alcance 2' WHEN fuente='Residuos' THEN 'Alcance 3' ELSE COALESCE(alcance,'') END")
        where.append(f"({_case}) = %s"); params.append(alcance)

    where_sql = " AND ".join(where)
    conn = get_db()
    df = pd.read_sql_query(
        f"SELECT fecha, area, alcance, fuente, categoria, actividad, cantidad, unidad, factor, emision "
        f"FROM registros WHERE {where_sql} ORDER BY fecha DESC",
        conn, params=params)
    conn.close()
    
    if df.empty:
        flash(f"La empresa no tiene registros para el periodo seleccionado ({anio}).", "warning")
        return redirect(request.referrer)
        
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        nombre_hoja = f'Auditoria_{anio}'
        df.to_excel(writer, sheet_name=nombre_hoja, index=False)
        worksheet = writer.sheets[nombre_hoja]
        for columna in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
            worksheet.column_dimensions[columna].width = 18
            
    output.seek(0)
    nombre_archivo = f"Reporte_Auditoria_{nombre_empresa.replace(' ', '_')}_{anio}.xlsx"
    return send_file(output, download_name=nombre_archivo, as_attachment=True)


@app.route("/admin/exportar_todo")
def admin_exportar_todo():
    if session.get('es_admin') != 1:
        return redirect("/")

    anio           = request.args.get('anio', 'Todos')
    empresa_filtro = request.args.get('empresa', '').strip()
    fuente_filtro  = request.args.get('fuente', '').strip()
    alcance_filtro = request.args.get('alcance', '').strip()

    where = ["1=1"]
    params = []
    if anio != 'Todos':
        where.append("SUBSTRING(fecha::text,1,4) = %s"); params.append(anio)
    if empresa_filtro:
        where.append("empresa = %s"); params.append(empresa_filtro)
    if fuente_filtro:
        where.append("fuente = %s"); params.append(fuente_filtro)
    if alcance_filtro:
        _case = ("CASE WHEN fuente IN ('Combustión Fija','Combustible Móvil','Combustión Estacionaria','Refrigerantes','Fugas de Refrigerantes') THEN 'Alcance 1'"
                 " WHEN fuente='Electricidad' THEN 'Alcance 2' WHEN fuente='Residuos' THEN 'Alcance 3' ELSE COALESCE(alcance,'') END")
        where.append(f"({_case}) = %s"); params.append(alcance_filtro)
    where_sql = " AND ".join(where)

    conn = get_db()
    df_total = pd.read_sql_query(
        f"SELECT empresa, fecha, area, alcance, fuente, categoria, actividad, cantidad, unidad, factor, emision "
        f"FROM registros WHERE {where_sql} ORDER BY empresa, fecha DESC",
        conn, params=params if params else None
    )
    empresas = pd.read_sql_query(
        "SELECT empresa FROM usuarios WHERE es_admin = 0 ORDER BY empresa",
        conn
    )['empresa'].tolist()
    conn.close()

    if df_total.empty:
        flash("No hay registros en la plataforma para exportar.", "warning")
        return redirect(url_for('admin_dashboard'))

    df_total.rename(columns={
        'empresa': 'Empresa', 'fecha': 'Fecha', 'area': 'Área',
        'alcance': 'Alcance', 'fuente': 'Fuente', 'categoria': 'Combustible/Categoría',
        'actividad': 'Uso', 'cantidad': 'Cantidad', 'unidad': 'Unidad',
        'factor': 'Factor', 'emision': 'Emisiones (kg CO2)'
    }, inplace=True)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_total.to_excel(writer, sheet_name='Todos', index=False)
        ws = writer.sheets['Todos']
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18

        for emp in empresas:
            df_emp = df_total[df_total['Empresa'] == emp]
            if df_emp.empty:
                continue
            nombre_hoja = emp[:31].replace('/', '-').replace('\\', '-').replace('*', '').replace('?', '').replace('[', '').replace(']', '').replace(':', '')
            df_emp.to_excel(writer, sheet_name=nombre_hoja, index=False)
            ws_emp = writer.sheets[nombre_hoja]
            for col in ws_emp.columns:
                ws_emp.column_dimensions[col[0].column_letter].width = 18

    output.seek(0)
    nombre_archivo = f"GreenTrack_Exportacion_Total_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(output, download_name=nombre_archivo, as_attachment=True)


# ================= OTROS MÓDULOS (Vehículos) =================
@app.route("/vehiculos")
def vehiculos():
    if 'user_id' not in session: return redirect("/")
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM vehiculos WHERE empresa = %s ORDER BY patente", (session.get('empresa'),))
    vehiculos_data = cursor.fetchall()
    conn.close()
    return render_template("vehiculos.html", vehiculos_data=vehiculos_data)

@app.route("/api/vehiculos", methods=["GET", "POST"])
def api_vehiculos():
    if 'user_id' not in session: return jsonify({"success": False, "message": "No autorizado"}), 401
    empresa = session.get('empresa')
    if request.method == "GET":
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("SELECT id, patente, tipo, marca, modelo, anio FROM vehiculos WHERE empresa = %s ORDER BY patente", (empresa,))
        vehiculos = cursor.fetchall()
        conn.close()
        return jsonify([{"id": v[0], "patente": v[1], "tipo": v[2], "marca": v[3], "modelo": v[4], "anio": v[5]} for v in vehiculos])
    elif request.method == "POST":
        data = request.get_json()
        if not data.get('patente'): return jsonify({"success": False, "message": "Patente obligatoria"}), 400
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("SELECT id FROM vehiculos WHERE empresa = %s AND patente = %s", (empresa, data['patente'].upper()))
        if cursor.fetchone():
            conn.close()
            return jsonify({"success": False, "message": "Patente ya registrada"}), 400
        cursor.execute("""
            INSERT INTO vehiculos (empresa, patente, tipo, marca, modelo, anio, fecha_registro)
            VALUES (%s, %s, %s, %s, %s, %s, %s) RETURNING id;
        """, (empresa, data['patente'].upper(), data.get('tipo'), data.get('marca'), data.get('modelo'), data.get('anio'), datetime.now().strftime("%Y-%m-%d %H:%M")))
        vehiculo_id = cursor.fetchone()[0]
        conn.commit()
        conn.close()
        return jsonify({"success": True, "message": "Vehículo guardado", "vehiculo": {"id": vehiculo_id, "patente": data['patente'].upper()}})

@app.route("/vehiculos/eliminar/<int:id>")
def eliminar_vehiculo(id):
    if 'user_id' not in session: return redirect("/")
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM vehiculos WHERE id = %s AND empresa = %s", (id, session.get('empresa')))
    conn.commit()
    conn.close()
    flash("Vehículo eliminado correctamente", "info")
    return redirect(url_for('vehiculos'))

@app.route("/combustible/movil/eliminar/<int:id>")
def eliminar_registro_combustible(id):
    if 'user_id' not in session: return redirect("/")
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM combustible_movil WHERE id = %s AND empresa = %s", (id, session.get('empresa')))
    conn.commit()
    conn.close()
    flash("Registro de combustible eliminado", "info")
    return redirect(url_for('combustible_movil'))

@app.route("/combustible/movil")
def combustible_movil():
    if 'user_id' not in session: return redirect("/")
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM vehiculos WHERE empresa = %s", (session.get('empresa'),))
    vehiculos = cursor.fetchall()
    conn.close()
    return render_template("combustible_movil.html", vehiculos=vehiculos)

@app.route("/api/combustible/movil", methods=["POST"])
def api_combustible_movil():
    if 'user_id' not in session: return jsonify({"success": False, "message": "No autorizado"}), 401
    empresa = session.get('empresa')
    data = request.get_json()
    if not data or not data.get('registros'):
        return jsonify({"success": False, "message": "No se recibieron registros"}), 400

    factores_movil = {
        'diesel': 2.68, 'bencina': 2.31, 'gas_natural': 2.02,
        'glp': 1.61, 'electricidad': 0.233, 'otro': 2.5
    }

    conn = get_db()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    guardados = 0
    errores = []

    for i, registro in enumerate(data.get('registros', [])):
        try:
            vehiculo_id = int(registro['vehiculo_id'])
            periodo = registro['periodo']  # YYYY-MM
            if not periodo:
                errores.append(f"Registro {i+1}: periodo es obligatorio")
                continue
            cantidad = float(registro['cantidad'])
            if cantidad <= 0:
                errores.append(f"Registro {i+1}: cantidad debe ser mayor a 0")
                continue

            combustible = registro['combustible']
            unidad = registro['unidad']
            costo = float(registro.get('costo') or 0)
            fecha = f"{periodo}-01"
            factor = factores_movil.get(combustible.lower(), 2.5)
            emision = round(cantidad * factor, 4)

            # Obtener patente del vehículo para registros
            cursor.execute("SELECT patente, tipo FROM vehiculos WHERE id = %s AND empresa = %s", (vehiculo_id, empresa))
            vehiculo = cursor.fetchone()
            patente = vehiculo['patente'] if vehiculo else f"Vehículo {vehiculo_id}"
            tipo_v = vehiculo['tipo'] if vehiculo else ''
            actividad = f"{patente} ({tipo_v})" if tipo_v else patente

            # 1. Insertar en registros (tabla maestra de emisiones)
            cursor.execute("""
                INSERT INTO registros (fecha, empresa, area, alcance, fuente, categoria, actividad, unidad, cantidad, factor, emision)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (fecha, empresa, 'Flota Vehicular', 'Alcance 1',
                  'Combustible Móvil', combustible, actividad,
                  unidad, cantidad, factor, emision))

            # 2. Insertar en combustible_movil (para tracking por vehículo)
            cursor.execute("""
                INSERT INTO combustible_movil (empresa, vehiculo_id, periodo, combustible, cantidad, unidad, costo, fecha_registro)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """, (empresa, vehiculo_id, periodo, combustible, cantidad, unidad, costo, fecha))

            guardados += 1
        except Exception as e:
            conn.rollback()
            errores.append(f"Registro {i+1}: {str(e)}")

    if guardados > 0:
        conn.commit()
    conn.close()
    if guardados == 0 and errores:
        return jsonify({"success": False, "message": "No se guardó ningún registro. Errores: " + "; ".join(errores)})
    return jsonify({"success": True, "guardados": guardados, "errores": errores})

@app.route("/combustible/fijo")
def combustible_fijo():
    if 'user_id' not in session: return redirect("/")
    return render_template("combustible_fijo.html")

@app.route("/api/combustible/fijo", methods=["POST"])
def api_combustible_fijo():
    if 'user_id' not in session: return redirect("/")
    empresa = session.get('empresa')
    combustible = request.form.get('combustible', '')
    unidad_consumo = request.form.get('unidad_consumo', '')
    unidad = request.form.get('unidad', 'litros')
    cantidad = float(request.form.get('cantidad', 0) or 0)
    costo = float(request.form.get('costo', 0) or 0)
    periodo = request.form.get('periodo', '')
    uso_final = request.form.get('uso_final', '')

    factores_fijo = {
        'diesel': 2.68, 'petroleo': 2.96, 'gas_natural': 2.02,
        'carbon': 2.42, 'gas_lp': 1.61, 'otro': 2.5
    }
    factor = factores_fijo.get(combustible, 2.5)
    emision = cantidad * factor

    fecha = (periodo + '-01') if periodo else datetime.now().strftime('%Y-%m-%d')

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO registros (fecha, empresa, fuente, categoria, actividad, identificador, unidad, cantidad, costo, factor, emision, alcance)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """, (fecha, empresa, 'Combustión Fija', combustible, uso_final or 'Combustión estacionaria',
          unidad_consumo, unidad, cantidad, costo, factor, emision, 'Alcance 1'))
    conn.commit()
    conn.close()
    flash("Registro de combustible fijo guardado correctamente.", "success")
    return redirect(url_for('combustible_fijo'))

@app.route("/mis_datos")
def mis_datos():
    if 'user_id' not in session: return redirect("/")
    conn = get_db()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cursor.execute("SELECT * FROM usuarios WHERE id = %s", (session['user_id'],))
    datos = cursor.fetchone()
    cursor.execute("SELECT DISTINCT EXTRACT(YEAR FROM fecha::date)::int AS anio FROM registros WHERE empresa = %s ORDER BY anio DESC", (session['empresa'],))
    anios = [r['anio'] for r in cursor.fetchall()]
    conn.close()
    return render_template("mis_datos.html", datos_usuario=datos, anios=anios)


@app.route("/mi_cuenta/guardar_preferencias", methods=["POST"])
def guardar_preferencias():
    if 'user_id' not in session: return redirect("/")
    anio_default = request.form.get("anio_default", "")
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("UPDATE usuarios SET anio_default = %s WHERE id = %s", (anio_default or None, session['user_id']))
    conn.commit()
    conn.close()
    flash("Preferencias guardadas correctamente.", "success")
    return redirect(url_for('mis_datos'))

@app.route("/cambiar_password", methods=["POST"])
def cambiar_password():
    if 'user_id' not in session:
        return redirect("/")
    actual = request.form.get("password_actual")
    nueva = request.form.get("password_nueva")
    confirmar = request.form.get("password_confirmar")
    if nueva != confirmar:
        flash("Las contraseñas nuevas no coinciden.", "error")
        return redirect(url_for('mis_datos'))
    if len(nueva) < 6:
        flash("La contraseña debe tener al menos 6 caracteres.", "error")
        return redirect(url_for('mis_datos'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT password FROM usuarios WHERE id = %s", (session['user_id'],))
    row = cursor.fetchone()
    if not row or not verify_password(actual, row[0]):
        conn.close()
        flash("La contraseña actual es incorrecta.", "error")
        return redirect(url_for('mis_datos'))
    cursor.execute("UPDATE usuarios SET password = %s WHERE id = %s", (hash_password(nueva), session['user_id']))
    conn.commit()
    conn.close()
    flash("Contraseña actualizada correctamente.", "success")
    return redirect(url_for('mis_datos'))

@app.route("/configuracion_sistema")
def configuracion_sistema(): return redirect(url_for('mis_datos'))

@app.route("/configuracion")
def configuracion(): return redirect(url_for('mis_datos'))

@app.route("/agua/registro", methods=["GET", "POST"])
def agua_registro(): return render_template("agua_registro.html")
@app.route("/agua/reporte")
def agua_reporte(): return render_template("agua_reporte.html")
@app.route("/residuos/registro", methods=["GET", "POST"])
def residuos_registro(): return render_template("residuos_registro.html")
@app.route("/residuos/reporte")
def residuos_reporte(): return render_template("residuos_reporte.html")
@app.route("/alcance_3")
def alcance_3(): return redirect(url_for('formulario_residuos'))

# ================= ELIMINACIÓN UNIVERSAL =================
@app.route('/eliminar_cualquier_registro/<tipo>/<int:id>')
def eliminar_cualquier_registro(tipo, id):
    if 'user_id' not in session: return redirect("/")
    empresa = session.get('empresa')
    conn = get_db()
    cursor = conn.cursor()
    try:
        if tipo == 'vehiculo':
            cursor.execute("DELETE FROM combustible_movil WHERE id = %s AND empresa = %s", (id, empresa))
        else:
            cursor.execute("DELETE FROM registros WHERE id = %s AND empresa = %s", (id, empresa))
        conn.commit()
        flash("Registro eliminado con éxito.", "success")
    except Exception as e:
        flash(f"Error al eliminar: {e}", "danger")
    finally:
        conn.close()
    # Magia: Te devuelve a la misma pantalla donde hiciste clic en el basurero
    return redirect(request.referrer or url_for('dashboard'))

# ================= HISTORIAL DE CARGAS PDF =================
@app.route("/residuos/pdf_historial")
def pdf_historial():
    if 'user_id' not in session: return redirect("/")
    empresa = session.get('empresa')
    conn = get_db()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cursor.execute("""
        SELECT id, fecha_subida, nombre_archivo, tipo, registros_generados, sin_factor
        FROM pdf_uploads WHERE empresa = %s ORDER BY fecha_subida DESC
    """, (empresa,))
    cargas = [dict(r) for r in cursor.fetchall()]
    conn.close()
    return render_template("pdf_historial.html", cargas=cargas)


# ================= HISTORIAL PDF GLOBAL (ADMIN) =================
@app.route("/admin/pdf_historial")
def admin_pdf_historial():
    if session.get('es_admin') != 1: return redirect("/")
    POR_PAGINA = 50
    pagina = max(1, int(request.args.get('pagina', 1)))
    offset = (pagina - 1) * POR_PAGINA
    conn = get_db()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cursor.execute("SELECT COUNT(*) FROM pdf_uploads")
    total = cursor.fetchone()[0]
    total_paginas = max(1, (total + POR_PAGINA - 1) // POR_PAGINA)
    cursor.execute("""
        SELECT empresa, fecha_subida, nombre_archivo, tipo, registros_generados, sin_factor
        FROM pdf_uploads ORDER BY fecha_subida DESC LIMIT %s OFFSET %s
    """, (POR_PAGINA, offset))
    cargas = [dict(r) for r in cursor.fetchall()]
    conn.close()
    return render_template("admin_pdf_historial.html", cargas=cargas,
                           pagina=pagina, total_paginas=total_paginas, total=total)


# ================= VALIDACIÓN DE PDFs (ADMIN) =================
@app.route("/admin/pendientes")
def admin_pendientes():
    if session.get('es_admin') != 1: return redirect("/")
    conn = get_db()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cursor.execute("SELECT * FROM pending_pdf_uploads ORDER BY CASE estado WHEN 'pendiente' THEN 0 ELSE 1 END, fecha_subida DESC")
    envios = [dict(r) for r in cursor.fetchall()]
    conn.close()
    for e in envios:
        try:
            e['filas'] = json.loads(e['datos_json'] or '[]')
        except Exception:
            e['filas'] = []
    return render_template("admin_pendientes.html", envios=envios)

@app.route("/admin/aprobar_pdf/<int:id>", methods=["POST"])
def admin_aprobar_pdf(id):
    if session.get('es_admin') != 1: return redirect("/")
    from datetime import datetime as dt
    conn = get_db()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cursor.execute("SELECT * FROM pending_pdf_uploads WHERE id = %s", (id,))
    pending = cursor.fetchone()
    if not pending:
        flash("Envío no encontrado.", "danger")
        conn.close()
        return redirect(url_for('admin_pendientes'))
    try:
        filas = json.loads(pending['datos_json'] or '[]')
        empresa = pending['empresa']
        for fila in filas:
            cursor.execute("""
                INSERT INTO registros (fecha, empresa, area, alcance, fuente, categoria, actividad, unidad, cantidad, factor, emision)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (fila['fecha'], empresa,
                  (fila.get('destino') or 'Operaciones')[:50],
                  'Alcance 3', 'Residuos', fila['categoria'], fila.get('tratamiento', ''),
                  'kg', fila['cantidad'], fila['factor'], fila['emision']))
        con_factor = sum(1 for f in filas if f['factor'] > 0)
        sin_factor = len(filas) - con_factor
        cursor.execute("""
            INSERT INTO pdf_uploads (empresa, fecha_subida, nombre_archivo, tipo, registros_generados, sin_factor)
            VALUES (%s, %s, %s, %s, %s, %s)
        """, (empresa, dt.now().strftime("%Y-%m-%d %H:%M"), pending['nombre_archivo'], pending['tipo'], len(filas), sin_factor))
        cursor.execute("""
            UPDATE pending_pdf_uploads SET estado = 'aprobado', fecha_revision = %s WHERE id = %s
        """, (dt.now().strftime("%Y-%m-%d %H:%M"), id))
        conn.commit()
        flash(f"Aprobado: {len(filas)} registros guardados para {empresa}.", "success")
    except Exception as e:
        conn.rollback()
        flash(f"Error al aprobar: {str(e)}", "danger")
    finally:
        conn.close()
    return redirect(url_for('admin_pendientes'))

@app.route("/admin/rechazar_pdf/<int:id>", methods=["POST"])
def admin_rechazar_pdf(id):
    if session.get('es_admin') != 1: return redirect("/")
    from datetime import datetime as dt
    motivo = request.form.get('motivo', '').strip() or 'Sin motivo especificado'
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        UPDATE pending_pdf_uploads SET estado = 'rechazado', motivo_rechazo = %s, fecha_revision = %s WHERE id = %s
    """, (motivo, dt.now().strftime("%Y-%m-%d %H:%M"), id))
    conn.commit()
    conn.close()
    flash("Envío rechazado.", "warning")
    return redirect(url_for('admin_pendientes'))

# ================= MIS ENVÍOS (USUARIO) =================
@app.route("/mis_envios")
def mis_envios():
    if 'user_id' not in session or session.get('es_admin') == 1: return redirect("/")
    empresa = session.get('empresa')
    conn = get_db()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cursor.execute("SELECT * FROM pending_pdf_uploads WHERE empresa = %s ORDER BY fecha_subida DESC", (empresa,))
    envios = [dict(r) for r in cursor.fetchall()]
    conn.close()
    for e in envios:
        try:
            e['filas'] = json.loads(e['datos_json'] or '[]')
        except Exception:
            e['filas'] = []
    return render_template("mis_envios.html", envios=envios)


# ================= API GHG GLOBAL (ADMIN) =================
@app.route("/api/admin/emisiones-por-empresa")
def api_admin_emisiones_por_empresa():
    if session.get('es_admin') != 1:
        return jsonify({"error": "No autorizado"}), 401

    conn = None
    try:
        anio = request.args.get('anio')
        anio = anio.strip() if anio else None
        filtro_anio = "WHERE SUBSTRING(fecha::text,1,4) = %s" if anio else ""

        query = f"""
            WITH base AS (
                SELECT
                    empresa,
                    SUBSTRING(fecha::text,1,4) AS anio,
                    CASE
                        WHEN fuente IN ('Combustión Fija','Combustible Móvil','Combustión Estacionaria','Refrigerantes','Fugas de Refrigerantes') THEN 'Alcance 1'
                        WHEN fuente = 'Electricidad' THEN 'Alcance 2'
                        WHEN fuente = 'Residuos' THEN 'Alcance 3'
                        ELSE COALESCE(alcance,'')
                    END AS alcance_calc,
                    COALESCE(emision, 0) AS emision,
                    COALESCE(emision_ubicacion, 0) AS emision_ubicacion
                FROM registros
                {filtro_anio}
            )
            SELECT
                empresa,
                anio,
                ROUND(COALESCE(SUM(CASE WHEN alcance_calc = 'Alcance 1' THEN emision ELSE 0 END), 0)::numeric, 2) AS alcance_1_kgco2e,
                ROUND(COALESCE(SUM(CASE WHEN alcance_calc = 'Alcance 2' THEN emision ELSE 0 END), 0)::numeric, 2) AS alcance_2_mercado_kgco2e,
                ROUND(COALESCE(SUM(CASE WHEN alcance_calc = 'Alcance 2' THEN emision_ubicacion ELSE 0 END), 0)::numeric, 2) AS alcance_2_ubicacion_kgco2e,
                ROUND(COALESCE(SUM(CASE WHEN alcance_calc = 'Alcance 3' THEN emision ELSE 0 END), 0)::numeric, 2) AS alcance_3_kgco2e,
                ROUND((
                    COALESCE(SUM(CASE WHEN alcance_calc = 'Alcance 1' THEN emision ELSE 0 END), 0)
                    + COALESCE(SUM(CASE WHEN alcance_calc = 'Alcance 2' THEN emision ELSE 0 END), 0)
                    + COALESCE(SUM(CASE WHEN alcance_calc = 'Alcance 3' THEN emision ELSE 0 END), 0)
                )::numeric, 2) AS total_ghg_kgco2e
            FROM base
            GROUP BY empresa, anio
            ORDER BY empresa, anio
        """

        conn = get_db()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        if anio:
            cursor.execute(query, (anio,))
        else:
            cursor.execute(query)

        filas = cursor.fetchall()
        respuesta = []
        for fila in filas:
            respuesta.append({
                "empresa": fila["empresa"],
                "anio": str(fila["anio"] or ""),
                "alcance_1_kgco2e": round(float(fila["alcance_1_kgco2e"] or 0), 2),
                "alcance_2_mercado_kgco2e": round(float(fila["alcance_2_mercado_kgco2e"] or 0), 2),
                "alcance_2_ubicacion_kgco2e": round(float(fila["alcance_2_ubicacion_kgco2e"] or 0), 2),
                "alcance_3_kgco2e": round(float(fila["alcance_3_kgco2e"] or 0), 2),
                "total_ghg_kgco2e": round(float(fila["total_ghg_kgco2e"] or 0), 2),
            })

        return jsonify(respuesta)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            conn.close()


# ================= EXPORTACIÓN GHG GLOBAL (ADMIN) =================
@app.route("/admin/exportar_ghg_global")
def admin_exportar_ghg_global():
    if session.get('es_admin') != 1: return redirect("/")
    anio = request.args.get('anio', str(datetime.now().year))

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT empresa FROM usuarios WHERE es_admin = 0 ORDER BY empresa")
    empresas = [r[0] for r in cursor.fetchall()]

    filtro = " AND SUBSTRING(fecha::text,1,4) = %s" if anio != 'Todos' else ""
    resumen_global = []
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for empresa in empresas:
            p1 = (empresa, anio) if anio != 'Todos' else (empresa,)

            _case_alc = """
                CASE
                    WHEN fuente IN ('Combustión Fija','Combustible Móvil','Combustión Estacionaria','Refrigerantes','Fugas de Refrigerantes') THEN 'Alcance 1'
                    WHEN fuente = 'Electricidad' THEN 'Alcance 2'
                    WHEN fuente = 'Residuos' THEN 'Alcance 3'
                    ELSE COALESCE(alcance,'')
                END
            """
            df_a1 = pd.read_sql_query(
                f"SELECT fecha,area,fuente,categoria,actividad,unidad,cantidad,factor,emision "
                f"FROM registros WHERE empresa=%s AND ({_case_alc})='Alcance 1'{filtro} ORDER BY fecha",
                conn, params=p1)
            df_a2 = pd.read_sql_query(
                f"SELECT fecha,area,fuente,categoria,actividad,unidad,cantidad,factor,emision,"
                f"COALESCE(emision_ubicacion,0) as emision_ubicacion "
                f"FROM registros WHERE empresa=%s AND ({_case_alc})='Alcance 2'{filtro} ORDER BY fecha",
                conn, params=p1)
            df_a3 = pd.read_sql_query(
                f"SELECT fecha,area,fuente,categoria,actividad,unidad,cantidad,factor,emision "
                f"FROM registros WHERE empresa=%s AND ({_case_alc})='Alcance 3'{filtro} ORDER BY fecha",
                conn, params=p1)

            total_a1 = float(df_a1['emision'].sum()) if not df_a1.empty else 0.0
            total_a2 = float(df_a2['emision'].sum()) if not df_a2.empty else 0.0
            total_a2_ub = float(df_a2['emision_ubicacion'].sum()) if not df_a2.empty else 0.0
            total_a3 = float(df_a3['emision'].sum()) if not df_a3.empty else 0.0
            resumen_global.append({
                "Empresa": empresa,
                "Alcance 1 (kg CO₂e)": round(total_a1, 2),
                "Alcance 2 mercado (kg CO₂e)": round(total_a2, 2),
                "Alcance 2 ubicación (kg CO₂e)": round(total_a2_ub, 2),
                "Alcance 3 (kg CO₂e)": round(total_a3, 2),
                "TOTAL GHG (kg CO₂e)": round(total_a1 + total_a2 + total_a3, 2),
            })

            if df_a1.empty and df_a2.empty and df_a3.empty:
                continue

            col_map = {
                'fecha': 'Fecha', 'area': 'Área', 'fuente': 'Fuente',
                'categoria': 'Categoría', 'actividad': 'Actividad',
                'unidad': 'Unidad', 'cantidad': 'Cantidad',
                'factor': 'Factor (kg CO₂/u)', 'emision': 'Emisión (kg CO₂e)',
                'emision_ubicacion': 'Emis. ubicación (kg CO₂e)'
            }
            dfs_empresa = []
            for alcance_name, df_alc in [('Alcance 1', df_a1), ('Alcance 2', df_a2), ('Alcance 3', df_a3)]:
                if not df_alc.empty:
                    df_copy = df_alc.copy()
                    df_copy.insert(0, 'Alcance', alcance_name)
                    df_copy.rename(columns={k: v for k, v in col_map.items() if k in df_copy.columns}, inplace=True)
                    dfs_empresa.append(df_copy)
            if dfs_empresa:
                nombre_hoja = empresa[:28].replace('/', '-').replace('\\', '-').replace('*', '').replace('?', '').replace('[', '').replace(']', '').replace(':', '')
                pd.concat(dfs_empresa, ignore_index=True).to_excel(writer, sheet_name=nombre_hoja, index=False)

        df_resumen = pd.DataFrame(resumen_global)
        df_resumen.to_excel(writer, sheet_name='Resumen Global', index=False)

        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill("solid", fgColor="064E3B")
        header_font = Font(bold=True, color="FFFFFF")
        total_fill = PatternFill("solid", fgColor="D1FAE5")
        total_font = Font(bold=True)
        for sheet_name, ws in writer.sheets.items():
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = 22
            if sheet_name == 'Resumen Global':
                for row_idx in range(2, ws.max_row + 1):
                    for cell in ws[row_idx]:
                        if cell.column_letter in ('F',):
                            cell.fill = total_fill
                            cell.font = total_font

    conn.close()
    output.seek(0)
    nombre = f"GHG_Global_{anio}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(output, download_name=nombre, as_attachment=True)


# ================= EXPORTACIÓN GHG PROTOCOL =================
@app.route("/exportar_ghg")
def exportar_ghg():
    if 'user_id' not in session: return redirect("/")
    empresa = session.get('empresa')
    anio = request.args.get('anio', str(datetime.now().year))

    conn = get_db()

    filtro = " AND SUBSTRING(fecha::text,1,4) = %s" if anio != 'Todos' else ""
    p1 = (empresa, anio) if anio != 'Todos' else (empresa,)

    _case_alcance = """
        CASE
            WHEN fuente IN ('Combustión Fija','Combustible Móvil','Combustión Estacionaria','Refrigerantes','Fugas de Refrigerantes') THEN 'Alcance 1'
            WHEN fuente = 'Electricidad' THEN 'Alcance 2'
            WHEN fuente = 'Residuos' THEN 'Alcance 3'
            ELSE COALESCE(alcance,'')
        END
    """
    df_a1 = pd.read_sql_query(
        f"SELECT fecha,area,fuente,categoria,actividad,unidad,cantidad,factor,emision "
        f"FROM registros WHERE empresa=%s AND ({_case_alcance})='Alcance 1'{filtro} ORDER BY fecha",
        conn, params=p1)

    df_a2 = pd.read_sql_query(
        f"SELECT fecha,area,fuente,categoria,actividad,unidad,cantidad,factor,emision,"
        f"COALESCE(emision_ubicacion,0) as emision_ubicacion,origen_energia,tiene_irec,COALESCE(sistema,'') as sistema "
        f"FROM registros WHERE empresa=%s AND ({_case_alcance})='Alcance 2'{filtro} ORDER BY fecha",
        conn, params=p1)

    df_a3 = pd.read_sql_query(
        f"SELECT fecha,area,fuente,categoria,actividad,unidad,cantidad,factor,emision "
        f"FROM registros WHERE empresa=%s AND ({_case_alcance})='Alcance 3'{filtro} ORDER BY fecha",
        conn, params=p1)

    conn.close()

    total_a1 = float(df_a1['emision'].sum()) if not df_a1.empty else 0.0
    total_a2_mercado = float(df_a2['emision'].sum()) if not df_a2.empty else 0.0
    total_a2_ubicacion = float(df_a2['emision_ubicacion'].sum()) if not df_a2.empty else 0.0
    total_a3 = float(df_a3['emision'].sum()) if not df_a3.empty else 0.0

    df_resumen = pd.DataFrame([
        {"Alcance": "Alcance 1 — Emisiones directas", "Total kg CO₂e": round(total_a1, 4)},
        {"Alcance": "Alcance 2 (mercado) — Electricidad comprada", "Total kg CO₂e": round(total_a2_mercado, 4)},
        {"Alcance": "Alcance 2 (ubicación) — Electricidad comprada", "Total kg CO₂e": round(total_a2_ubicacion, 4)},
        {"Alcance": "Alcance 3 — Residuos y otras indirectas", "Total kg CO₂e": round(total_a3, 4)},
        {"Alcance": "TOTAL GHG", "Total kg CO₂e": round(total_a1 + total_a2_mercado + total_a3, 4)},
    ])

    col_map = {
        'fecha': 'Fecha', 'area': 'Área / Sucursal', 'fuente': 'Fuente',
        'categoria': 'Categoría / Combustible', 'actividad': 'Actividad / Uso',
        'unidad': 'Unidad', 'cantidad': 'Cantidad', 'factor': 'Factor (kg CO₂/u)',
        'emision': 'Emisión (kg CO₂e)', 'emision_ubicacion': 'Emisión ubicación (kg CO₂e)',
        'origen_energia': 'Origen Energía', 'tiene_irec': 'IREC', 'sistema': 'Sistema Eléctrico'
    }
    for df in [df_a1, df_a2, df_a3]:
        df.rename(columns={k: v for k, v in col_map.items() if k in df.columns}, inplace=True)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_resumen.to_excel(writer, sheet_name='Resumen GHG', index=False)
        if not df_a1.empty:
            df_a1.to_excel(writer, sheet_name='Alcance 1', index=False)
        if not df_a2.empty:
            df_a2.to_excel(writer, sheet_name='Alcance 2 Electricidad', index=False)
        if not df_a3.empty:
            df_a3.to_excel(writer, sheet_name='Alcance 3', index=False)

        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill("solid", fgColor="064E3B")
        header_font = Font(bold=True, color="FFFFFF")
        total_fill = PatternFill("solid", fgColor="D1FAE5")
        total_font = Font(bold=True)

        for sheet_name, ws in writer.sheets.items():
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = 22
            if sheet_name == 'Resumen GHG':
                last_row = ws.max_row
                for cell in ws[last_row]:
                    cell.fill = total_fill
                    cell.font = total_font

    output.seek(0)
    nombre = f"GHG_Protocol_{empresa.replace(' ','_')}_{anio}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(output, download_name=nombre, as_attachment=True)


# ─── TICKETS ──────────────────────────────────────────────────────────────────

@app.route('/tickets')
def tickets():
    if not session.get('user_id'):
        return redirect('/login')
    if session.get('es_admin') == 1:
        return redirect('/admin/tickets')
    empresa = session.get('empresa')
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute("SELECT * FROM tickets WHERE empresa = %s ORDER BY fecha_creacion DESC", (empresa,))
    mis_tickets = cur.fetchall()
    conn.close()
    return render_template('tickets.html', tickets=mis_tickets)


@app.route('/tickets/nuevo', methods=['POST'])
def tickets_nuevo():
    if not session.get('user_id') or session.get('es_admin') == 1:
        return redirect('/login')
    empresa = session.get('empresa')
    asunto = request.form.get('asunto', '').strip()
    descripcion = request.form.get('descripcion', '').strip()
    prioridad = request.form.get('prioridad', 'Normal')
    if not asunto or not descripcion:
        flash('El asunto y la descripción son obligatorios.', 'error')
        return redirect('/tickets')
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO tickets (empresa, asunto, descripcion, prioridad, estado, fecha_creacion)
        VALUES (%s, %s, %s, %s, 'Abierto', %s)
    """, (empresa, asunto, descripcion, prioridad, datetime.now().strftime('%Y-%m-%d %H:%M')))
    conn.commit()
    conn.close()
    flash('Ticket enviado correctamente. Te responderemos pronto.', 'success')
    return redirect('/tickets')


@app.route('/admin/tickets')
def admin_tickets():
    if session.get('es_admin') != 1:
        return redirect('/login')
    filtro_estado = request.args.get('estado', '')
    filtro_empresa = request.args.get('empresa', '')
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    where = []
    params = []
    if filtro_estado:
        where.append("estado = %s")
        params.append(filtro_estado)
    if filtro_empresa:
        where.append("empresa ILIKE %s")
        params.append(f'%{filtro_empresa}%')
    sql = "SELECT * FROM tickets"
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY fecha_creacion DESC"
    cur.execute(sql, params)
    all_tickets = cur.fetchall()
    cur.execute("SELECT estado, COUNT(*) FROM tickets GROUP BY estado")
    conteos = {row[0]: row[1] for row in cur.fetchall()}
    conn.close()
    return render_template('admin_tickets.html', tickets=all_tickets, conteos=conteos,
                           filtro_estado=filtro_estado, filtro_empresa=filtro_empresa)


@app.route('/admin/tickets/<int:ticket_id>/responder', methods=['POST'])
def admin_tickets_responder(ticket_id):
    if session.get('es_admin') != 1:
        return redirect('/login')
    respuesta = request.form.get('respuesta', '').strip()
    nuevo_estado = request.form.get('estado', 'Abierto')
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        UPDATE tickets SET respuesta = %s, estado = %s, fecha_respuesta = %s WHERE id = %s
    """, (respuesta, nuevo_estado, datetime.now().strftime('%Y-%m-%d %H:%M'), ticket_id))
    conn.commit()
    conn.close()
    flash('Respuesta enviada correctamente.', 'success')
    return redirect('/admin/tickets')


# Inicia la DB
init_db()

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
